import os
import io
import math
import json
from datetime import datetime, timedelta, timezone
from flask import (Flask, render_template, request, redirect, url_for, session,
                   jsonify, send_file, g, has_request_context)
from flask_admin import Admin, BaseView, expose, AdminIndexView
from flask_admin.contrib.sqla import ModelView
from sqlalchemy import inspect as sa_inspect, text, event
from werkzeug.security import generate_password_hash, check_password_hash

from models import (db, Camera, RentalBooking, PurchaseOrder, Sheet,
                    StoreCost, RepairLog, ActivityLog,
                    Accessory, AccessorySale, AccessoryDamage, Appointment, Lead,
                    Employee, Customer, Setting)
from seed import seed as seed_db, backfill_costs

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    HAS_OPENPYXL = True
except ImportError:                       # pragma: no cover
    HAS_OPENPYXL = False

# Camera categories for the Selling tab "type" dropdown.
CAMERA_CATEGORIES = ['Mirrorless', 'Compact', 'DSLR', 'Máy phim', 'Action Cam',
                     'Flycam', 'Chụp lấy liền', 'Khác']

# Where a buying customer came from (sale popup dropdown).
SALE_SOURCES = ['Facebook', 'Instagram', 'Threads', 'TikTok', 'Tại shop', 'Khác']

# Accessory / other-goods categories (Phụ kiện tab).
ACCESSORY_CATEGORIES = ['Lens', 'Pin', 'Thẻ nhớ', 'Sạc', 'Túi/Bao', 'Dây đeo',
                        'Tripod', 'Filter', 'Khác']

# Column order of the Excel import template (mirrors the shop's spreadsheet).
IMPORT_COLUMNS = ['Tên máy', 'Màu', 'Dòng máy', 'Nơi (N/V)', 'Phụ kiện',
                  'Ngày về (YYYY-MM-DD)', 'Ghi chú', 'Giá nhập', 'Giá bán', 'Trạng thái']

# state text (Vietnamese / English) → internal sale_state
STATE_ALIASES = {
    'còn hàng': 'stock', 'con hang': 'stock', 'stock': 'stock', 'tồn': 'stock', '': 'stock',
    'đang xử lý': 'processing', 'dang xu ly': 'processing', 'processing': 'processing', 'xử lý': 'processing',
    'đã cọc': 'deposit', 'da coc': 'deposit', 'deposit': 'deposit', 'cọc': 'deposit',
    'đã bán': 'sold', 'da ban': 'sold', 'sold': 'sold', 'bán': 'sold',
    'cần sửa': 'fixing', 'can sua': 'fixing', 'fixing': 'fixing', 'sửa': 'fixing', 'hỏng': 'fixing',
    'không sửa được': 'unfixable', 'khong sua duoc': 'unfixable', 'unfixable': 'unfixable',
}

# Vietnamese labels offered in the Excel "Trạng thái" dropdown (order matters for the sheet)
STATE_CHOICES = ['Còn hàng', 'Đang xử lý', 'Đã cọc', 'Đã bán', 'Cần sửa', 'Không sửa được']


def normalize_origin(value):
    """Map a free-text origin cell to 'Nhật' / 'Việt' / ''."""
    v = (str(value or '')).strip().lower()
    if v in ('n', 'nhật', 'nhat', 'japan', 'jp', 'nhật bản', 'nhat ban'):
        return 'Nhật'
    if v in ('v', 'việt', 'viet', 'vietnam', 'vn', 'việt nam', 'viet nam'):
        return 'Việt'
    return ''


def normalize_state(value):
    return STATE_ALIASES.get((str(value or '')).strip().lower(), 'stock')


def _to_int(value):
    """Coerce a spreadsheet cell (number / '1.200.000' / '1,200,000') to int."""
    if value is None or value == '':
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    digits = ''.join(ch for ch in str(value) if ch.isdigit())
    return int(digits) if digits else 0


def _cell_date(value):
    if isinstance(value, datetime):
        return value
    return parse_dt(str(value)) if value else None

basedir = os.path.abspath(os.path.dirname(__file__))
# Templates and static assets live in the sibling frontend/ folder.
frontend_dir = os.path.join(basedir, '..', 'frontend')
template_dir = os.path.join(frontend_dir, 'templates')
static_dir   = os.path.join(frontend_dir, 'static')

app = Flask(__name__, template_folder=template_dir,
            static_folder=static_dir, static_url_path='/static')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///camerashop.db'
# Session-signing key. MUST be set via env in production (a leaked/default key
# lets anyone forge an admin session cookie).
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-insecure-change-me')
app.config['TEMPLATES_AUTO_RELOAD'] = True   # pick up template edits without a restart
app.jinja_env.auto_reload = True

# ── Admin security ───────────────────────────────────────────────────────────
# The whole admin panel lives behind a secret, hard-to-guess URL prefix AND a
# login. Both are configurable via environment variables (see SECURITY.md).
ADMIN_URL_PREFIX = '/' + os.environ.get('ADMIN_URL_PREFIX', 'quanly-tintus').strip('/')
ADMIN_USERNAME   = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD_HASH = generate_password_hash(os.environ.get('ADMIN_PASSWORD', 'tintus@2026'))
LOGIN_PATH  = ADMIN_URL_PREFIX + '/login'
LOGOUT_PATH = ADMIN_URL_PREFIX + '/logout'

# Harden the session cookie. Set SESSION_COOKIE_SECURE=1 once you only serve
# the admin over HTTPS (e.g. behind Cloudflare) — leave 0 for plain-HTTP LAN access.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=(os.environ.get('SESSION_COOKIE_SECURE', '0') == '1'),
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8),
)

db.init_app(app)


# ══ Activity log: automatic capture + undo / redo ═══════════════════════════════
# Every admin DB mutation (insert / update / delete) is captured automatically via
# SQLAlchemy session events, grouped per request into one ActivityLog entry, so any
# action can be reversed (undo) and re-applied (redo) without touching each endpoint.

# Models whose row changes are tracked. ActivityLog itself is intentionally excluded
# (so writing the log never logs itself, and undo/redo can never target the log).
# Employee/Setting are intentionally NOT logged (snapshots would store password
# hashes, and undo could silently roll back credentials / config).
LOGGED_MODELS = [Camera, RentalBooking, Appointment, Lead, RepairLog,
                 StoreCost, Accessory, AccessorySale, AccessoryDamage, PurchaseOrder, Sheet,
                 Customer]
TABLE_MODELS = {m.__tablename__: m for m in LOGGED_MODELS}

# Vietnamese labels for auto-generated activity descriptions.
_ACT_TABLE_VI = {
    'camera': 'máy ảnh', 'rental_booking': 'đơn thuê', 'appointment': 'lịch hẹn',
    'lead': 'khách quan tâm', 'repair_log': 'sửa chữa', 'store_cost': 'chi phí tiệm',
    'accessory': 'phụ kiện', 'accessory_sale': 'bán phụ kiện',
    'accessory_damage': 'phụ kiện hư',
    'purchase_order': 'đơn mua', 'sheet': 'sổ tay', 'customer': 'khách hàng',
}
_ACT_VERB = {'insert': 'Thêm', 'update': 'Sửa', 'delete': 'Xoá'}


def _act_default(o):
    """JSON encoder: make datetime round-trippable."""
    if isinstance(o, datetime):
        return {'__dt__': o.isoformat()}
    raise TypeError(f'not serializable: {type(o)}')


def _act_hook(dct):
    if '__dt__' in dct:
        try:
            return datetime.fromisoformat(dct['__dt__'])
        except (ValueError, TypeError):
            return None
    return dct


def _act_dumps(x):
    return json.dumps(x, default=_act_default, ensure_ascii=False)


def _act_loads(s):
    return json.loads(s or '[]', object_hook=_act_hook)


def _act_snapshot(obj):
    """Full column snapshot of a mapped instance."""
    return {c.key: getattr(obj, c.key) for c in sa_inspect(obj).mapper.column_attrs}


def _act_active():
    """True when we should be capturing changes for the current request."""
    return has_request_context() and not getattr(g, '_act_suspend', False)


@event.listens_for(db.session, 'before_flush')
def _act_before_flush(session_, flush_context, instances):
    if not _act_active():
        return
    changes = getattr(g, '_act_changes', None)
    if changes is None:
        changes = g._act_changes = []
        g._act_new = []
    # Inserts: pk not assigned yet → stash the object, snapshot after flush.
    for obj in session_.new:
        if obj.__tablename__ in TABLE_MODELS:
            g._act_new.append(obj)
    # Deletes: object still fully loaded → snapshot now for later restore.
    for obj in session_.deleted:
        if obj.__tablename__ in TABLE_MODELS:
            changes.append({'action': 'delete', 'table': obj.__tablename__,
                            'pk': obj.id, 'before': _act_snapshot(obj), 'after': None})
    # Updates: record only the columns that actually changed (old → new).
    for obj in session_.dirty:
        if obj.__tablename__ not in TABLE_MODELS:
            continue
        st = sa_inspect(obj)
        before, after = {}, {}
        for attr in st.mapper.column_attrs:
            hist = st.attrs[attr.key].history
            if hist.has_changes():
                before[attr.key] = hist.deleted[0] if hist.deleted else None
                after[attr.key] = hist.added[0] if hist.added else None
        if after:
            changes.append({'action': 'update', 'table': obj.__tablename__,
                            'pk': obj.id, 'before': before, 'after': after})


@event.listens_for(db.session, 'after_flush')
def _act_after_flush(session_, flush_context):
    if not _act_active():
        return
    pending = getattr(g, '_act_new', None)
    if not pending:
        return
    changes = g._act_changes
    for obj in pending:
        changes.append({'action': 'insert', 'table': obj.__tablename__,
                        'pk': obj.id, 'before': None, 'after': _act_snapshot(obj)})
    g._act_new = []


def _act_title(change):
    snap = change.get('after') or change.get('before') or {}
    for k in ('name', 'customer_name', 'note', 'category', 'label'):
        if snap.get(k):
            return str(snap[k])
    # An update snapshot only holds the changed columns → look up the live row's name.
    Model = TABLE_MODELS.get(change.get('table'))
    if Model and change.get('pk') is not None:
        obj = db.session.get(Model, change['pk'])
        if obj:
            for k in ('name', 'customer_name', 'note', 'category'):
                v = getattr(obj, k, None)
                if v:
                    return str(v)
    return f"#{change.get('pk')}"


def _act_auto_label(changes):
    """A short Vietnamese summary of a change set, e.g. 'Sửa máy ảnh: Canon G7X'."""
    primary = changes[0]
    for c in changes:                       # prefer the main entity over side-effects
        if c['table'] not in ('repair_log', 'accessory_sale'):
            primary = c
            break
    label = f"{_ACT_VERB.get(primary['action'], 'Thay đổi')} " \
            f"{_ACT_TABLE_VI.get(primary['table'], primary['table'])}: {_act_title(primary)}"
    extra = len(changes) - 1
    if extra > 0:
        label += f" (+{extra} thay đổi)"
    return label


@app.after_request
def _act_persist(response):
    """After each successful admin write, save the captured change set as one activity."""
    if not has_request_context():
        return response
    changes = getattr(g, '_act_changes', None)
    if not changes:
        return response
    # Only admin actions belong in the admin history; never log during undo/redo.
    if getattr(g, '_act_suspend', False) or not request.path.startswith(ADMIN_URL_PREFIX):
        return response
    if response.status_code >= 400:         # the write failed / was rejected
        return response
    try:
        g._act_suspend = True
        # A new action truncates the redo tail (standard linear undo/redo).
        ActivityLog.query.filter_by(undone=True).delete()
        label = getattr(g, '_act_label', '') or _act_auto_label(changes)
        who = session.get('staff_user')
        if who:
            label += f' · {who}'
        db.session.add(ActivityLog(
            label=label[:300],
            changes_json=_act_dumps(changes), created_at=datetime.now()))
        db.session.commit()
    except Exception as exc:                # never let logging break the real response
        db.session.rollback()
        app.logger.warning(f'[activity] persist failed: {exc}')
    finally:
        g._act_suspend = False
        g._act_changes = []
    return response


def _act_restore(obj, snap):
    for k, v in snap.items():
        setattr(obj, k, v)


def _act_reverse(changes):
    """Undo: walk the change set backwards, applying the inverse of each row change."""
    for c in reversed(changes):
        Model = TABLE_MODELS.get(c['table'])
        if not Model:
            continue
        if c['action'] == 'insert':                     # created → delete it
            obj = db.session.get(Model, c['pk'])
            if obj:
                db.session.delete(obj)
        elif c['action'] == 'delete':                   # deleted → recreate it
            if db.session.get(Model, c['pk']) is None:
                obj = Model()
                _act_restore(obj, c['before'])
                db.session.add(obj)
        elif c['action'] == 'update':                   # changed → set old values
            obj = db.session.get(Model, c['pk'])
            if obj:
                _act_restore(obj, c['before'])


def _act_apply(changes):
    """Redo: walk the change set forwards, re-applying each row change."""
    for c in changes:
        Model = TABLE_MODELS.get(c['table'])
        if not Model:
            continue
        if c['action'] == 'insert':                     # re-create the row
            if db.session.get(Model, c['pk']) is None:
                obj = Model()
                _act_restore(obj, c['after'])
                db.session.add(obj)
        elif c['action'] == 'delete':                   # re-delete the row
            obj = db.session.get(Model, c['pk'])
            if obj:
                db.session.delete(obj)
        elif c['action'] == 'update':                   # set new values again
            obj = db.session.get(Model, c['pk'])
            if obj:
                _act_restore(obj, c['after'])


def _act_detail_lines(changes):
    """Human-readable per-row lines for the expandable history detail."""
    lines = []
    for c in changes:
        noun = _ACT_TABLE_VI.get(c['table'], c['table'])
        verb = _ACT_VERB.get(c['action'], '?')
        if c['action'] == 'update':
            before, after = c.get('before') or {}, c.get('after') or {}
            parts = [f"{k}: {_act_fmt(before.get(k))} → {_act_fmt(after.get(k))}" for k in after]
            lines.append(f"{verb} {noun} #{c['pk']} — " + '; '.join(parts))
        else:
            lines.append(f"{verb} {noun} #{c['pk']}: {_act_title(c)}")
    return lines


def _act_fmt(v):
    if v is None or v == '':
        return '∅'
    if isinstance(v, datetime):
        return v.strftime('%d/%m/%Y %H:%M')
    s = str(v)
    return s if len(s) <= 40 else s[:37] + '…'


class RedirectIndexView(AdminIndexView):
    """The default Flask-Admin home is unused → send it straight to Cameras."""
    @expose('/')
    def index(self):
        return redirect(url_for('cameras.index'))


class StaffView(BaseView):
    """Base for panel views reachable by every logged-in staff member."""
    def is_accessible(self):
        return current_role() in ('admin', 'employee')

    def inaccessible_callback(self, name, **kwargs):
        if not current_role():
            return redirect(url_for('admin_login', next=request.path))
        # logged in but not allowed (employee hitting an admin view) → their home
        if request.accept_mimetypes.best == 'application/json' or request.method == 'POST':
            return _forbidden()
        return redirect(url_for('cameras.index'))


class AdminOnlyView(StaffView):
    """Base for panel views reserved for the admin role (finances, personnel…)."""
    def is_accessible(self):
        return is_admin()


class SecureModelView(ModelView):
    """Raw Flask-Admin model tables — admin only (they expose every column)."""
    def is_accessible(self):
        return is_admin()

    def inaccessible_callback(self, name, **kwargs):
        if not current_role():
            return redirect(url_for('admin_login', next=request.path))
        return redirect(url_for('cameras.index'))


admin = Admin(app, name='tintus.digicam Admin',
              index_view=RedirectIndexView(url=ADMIN_URL_PREFIX, endpoint='admin'))
admin.add_view(SecureModelView(Camera, db.session))
admin.add_view(SecureModelView(RentalBooking, db.session))
admin.add_view(SecureModelView(PurchaseOrder, db.session))


# ── Roles & current-user helpers ──────────────────────────────────────────────
# Two roles share the panel behind the same secret URL prefix:
#   'admin'    — everything (finances, personnel, system config, full CRUD)
#   'employee' — day-to-day operations only; never sees import costs / profit.

def current_role():
    role = session.get('role')
    if role in ('admin', 'employee'):
        return role
    return 'admin' if session.get('is_admin') else None   # legacy sessions


def is_admin():
    return current_role() == 'admin'


def current_staff():
    """Display name of whoever is logged in (used for attribution & dashboards)."""
    return session.get('staff_name') or session.get('staff_user') or ''


def current_staff_user():
    return session.get('staff_user') or ''


def _forbidden(msg='Bạn không có quyền thực hiện thao tác này.'):
    return jsonify({'ok': False, 'error': msg}), 403


# ── System settings (key/value, admin-configurable) ───────────────────────────
DEFAULT_SETTINGS = {
    'store_name':       'tintus.digicam',
    'store_address':    '',
    'store_phone':      '',
    'tax_rate':         '0',    # % VAT shown on receipts
    'max_discount_pct': '10',   # employee sale-price floor: listed price minus this %
    'receipt_header':   'tintus.digicam — Cho thuê & mua bán máy ảnh',
    'receipt_footer':   'Cảm ơn quý khách! Hẹn gặp lại.',
    'store_policy':     '',
}


def get_setting(key):
    row = db.session.get(Setting, key)
    return row.value if row and row.value is not None else DEFAULT_SETTINGS.get(key, '')


def all_settings():
    return {k: get_setting(k) for k in DEFAULT_SETTINGS}


def setting_float(key):
    try:
        return float(get_setting(key) or 0)
    except (TypeError, ValueError):
        return 0.0


# ── Gate every admin request behind the login ────────────────────────────────
@app.before_request
def _require_admin_login():
    p = request.path
    if not (p == ADMIN_URL_PREFIX or p.startswith(ADMIN_URL_PREFIX + '/')):
        return                                  # public site → no gate
    if p in (LOGIN_PATH, LOGOUT_PATH):
        return                                  # auth endpoints stay reachable
    if not current_role():
        return redirect(url_for('admin_login', next=p))


@app.route(LOGIN_PATH, methods=['GET', 'POST'])
def admin_login():
    if current_role():
        return redirect(url_for('cameras.index'))
    error = None
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''
        role = staff_user = staff_name = None
        # 1) the env-configured owner account (always an admin)
        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH, password):
            role, staff_user, staff_name = 'admin', username, 'Chủ tiệm'
        else:
            # 2) staff accounts managed in the Nhân viên view
            emp = Employee.query.filter(db.func.lower(Employee.username) == username.lower()).first()
            if emp and emp.active and check_password_hash(emp.password_hash, password):
                role = emp.role if emp.role in ('admin', 'employee') else 'employee'
                staff_user, staff_name = emp.username, emp.display_name or emp.username
        if role:
            session.clear()
            session['role'] = role
            session['is_admin'] = (role == 'admin')     # legacy flag, kept in sync
            session['staff_user'] = staff_user
            session['staff_name'] = staff_name
            session.permanent = True
            nxt = request.args.get('next') or url_for('cameras.index')
            if not nxt.startswith(ADMIN_URL_PREFIX):     # block open-redirects
                nxt = url_for('cameras.index')
            return redirect(nxt)
        error = 'Sai tài khoản hoặc mật khẩu (hoặc tài khoản đã bị khoá).'
    return render_template('admin_login.html', error=error, login_action=LOGIN_PATH)


@app.route(LOGOUT_PATH)
def admin_logout():
    session.clear()
    return redirect(LOGIN_PATH)


# ── Role info available to every admin template ───────────────────────────────
@app.context_processor
def inject_role():
    return {'is_admin': is_admin(), 'current_role': current_role(),
            'staff_name': current_staff(), 'staff_user': current_staff_user()}


# ── Camera sub-tab counts for the sidebar (Cho thuê / Để bán / Đã cọc / Đã bán) ─
@app.context_processor
def inject_nav_counts():
    """Feeds the expandable "Quản lý máy" sidebar group. Only touches the DB for
    logged-in admin requests, so public pages pay nothing."""
    if not current_role():
        return {}
    try:
        counts = {
            'rent':    Camera.query.filter_by(type='Rent').count(),
            'sale':    Camera.query.filter_by(type='Sale')
                             .filter(~Camera.sale_state.in_(('sold', 'deposit'))).count(),
            'deposit': Camera.query.filter_by(type='Sale', sale_state='deposit').count(),
            'sold':    Camera.query.filter_by(type='Sale', sale_state='sold').count(),
        }
    except Exception:
        counts = {'rent': 0, 'sale': 0, 'deposit': 0, 'sold': 0}
    return {'cam_tab_counts': counts}


# ── Jinja2 filter: 280000 → "280K" ─────────────────────────────────────────
@app.template_filter('vnd')
def vnd_filter(value):
    """Full VND with Vietnamese thousands separators: 280000 → '280.000' (— if empty)."""
    if not value:
        return '—'
    return '{:,.0f}'.format(int(value)).replace(',', '.')


# ── Jinja2 filter: 1200000 → "1.200.000" (Vietnamese thousands separators) ──
@app.template_filter('money')
def money_filter(value):
    return '{:,.0f}'.format(int(value or 0)).replace(',', '.')


# ── Global template context: cart badge count on every page ─────────────────
@app.context_processor
def inject_cart_count():
    return {'cart_count': sum(session.get('cart', {}).values())}


# ── Brand display config (used in templates) ────────────────────────────────
BRAND_CONFIG = {
    'Canon':   {'header': 'bg-red-700',   'hover': 'hover:bg-red-50',   'text': 'text-red-700',   'icon': 'bg-red-100',   'emoji': '📷'},
    'Fujifilm':{'header': 'bg-green-700', 'hover': 'hover:bg-green-50', 'text': 'text-green-700', 'icon': 'bg-green-100', 'emoji': '📷'},
    'Sony':    {'header': 'bg-blue-700',  'hover': 'hover:bg-blue-50',  'text': 'text-blue-700',  'icon': 'bg-blue-100',  'emoji': '📷'},
    'DJI':     {'header': 'bg-gray-700',  'hover': 'hover:bg-gray-50',  'text': 'text-gray-700',  'icon': 'bg-gray-100',  'emoji': '🎥'},
    'Lumix':   {'header': 'bg-indigo-700','hover': 'hover:bg-indigo-50','text': 'text-indigo-700','icon': 'bg-indigo-100','emoji': '📷'},
    'Casio':   {'header': 'bg-yellow-600','hover': 'hover:bg-yellow-50','text': 'text-yellow-700','icon': 'bg-yellow-100','emoji': '📷'},
    'Nikon':   {'header': 'bg-yellow-700','hover': 'hover:bg-yellow-50','text': 'text-yellow-700','icon': 'bg-yellow-100','emoji': '📷'},
}


# ── ROUTES ──────────────────────────────────────────────────────────────────

@app.route('/')
def home():
    featured = Camera.query.filter_by(type='Rent', featured=True).order_by(Camera.price.desc()).limit(8).all()
    return render_template('index.html', featured=featured)


@app.route('/lead', methods=['POST'])
def submit_lead():
    """Capture a homepage inquiry as a follow-up lead (no camera/date validation needed)."""
    name  = (request.form.get('customer_name') or '').strip()
    phone = (request.form.get('phone') or '').strip()
    if not (name or phone):                     # ignore empty submissions
        return redirect(url_for('home'))
    db.session.add(Lead(
        customer_name=name, phone=phone,
        notes=(request.form.get('notes') or '').strip(),
        start_date=parse_dt(request.form.get('start_date')),
        end_date=parse_dt(request.form.get('end_date')),
        source='Trang chủ'))
    db.session.commit()
    return redirect(url_for('home', sent=1) + '#dat-lich')


@app.route('/store')
def store():
    brand = request.args.get('brand', '').strip()
    keyword = request.args.get('q', '').strip()
    price = request.args.get('price', '').strip()
    sort = request.args.get('sort', '').strip()

    query = Camera.query.filter_by(type='Rent')
    if brand:
        query = query.filter(Camera.brand.ilike(brand))
    if keyword:
        like = f'%{keyword}%'
        query = query.filter(db.or_(Camera.name.ilike(like), Camera.brand.ilike(like)))
    if price == 'low':
        query = query.filter(Camera.price < 200000)
    elif price == 'mid':
        query = query.filter(Camera.price >= 200000, Camera.price <= 350000)
    elif price == 'high':
        query = query.filter(Camera.price > 350000)

    if sort == 'price_asc':
        query = query.order_by(Camera.price.asc())
    elif sort == 'price_desc':
        query = query.order_by(Camera.price.desc())
    elif sort == 'name':
        query = query.order_by(Camera.name.asc())
    else:
        query = query.order_by(Camera.brand, Camera.price.desc())

    cameras = query.all()
    brands = [r[0] for r in db.session.query(Camera.brand).filter_by(type='Rent').distinct().order_by(Camera.brand).all()]
    return render_template('store.html', cameras=cameras, current_brand=brand, brands=brands,
                           keyword=keyword, current_price=price, current_sort=sort)


@app.route('/mua')
def buy_store():
    """Public storefront for cameras that are for SALE (one inventory truth: variants)."""
    brand = request.args.get('brand', '').strip()
    keyword = request.args.get('q', '').strip()
    query = Camera.query.filter_by(type='Sale', sale_state='stock')   # only available units
    if brand:
        query = query.filter(Camera.brand.ilike(brand))
    if keyword:
        like = f'%{keyword}%'
        query = query.filter(db.or_(Camera.name.ilike(like), Camera.brand.ilike(like)))
    cameras = query.order_by(Camera.brand, Camera.price.desc()).all()
    brands = [r[0] for r in db.session.query(Camera.brand).filter_by(type='Sale', sale_state='stock').distinct().order_by(Camera.brand).all() if r[0]]
    return render_template('mua.html', cameras=cameras, current_brand=brand, brands=brands, keyword=keyword)


@app.route('/san-pham')
def san_pham():
    return redirect(url_for('store'))


@app.route('/bang-gia')
def bang_gia():
    cameras = Camera.query.filter_by(type='Rent').order_by(Camera.brand, Camera.price.desc()).all()
    # Group by brand preserving order
    grouped = {}
    for cam in cameras:
        grouped.setdefault(cam.brand, []).append(cam)
    return render_template('bang_gia.html', grouped=grouped, brand_config=BRAND_CONFIG)


@app.route('/bang-gia-cho-thue-may-anh-2saigon')
def bang_gia_alias():
    return redirect(url_for('bang_gia'))


@app.route('/product/<slug>')
def product_detail(slug):
    product = Camera.query.filter_by(slug=slug).first_or_404()
    related = (Camera.query
               .filter_by(brand=product.brand, type='Rent')
               .filter(Camera.id != product.id)
               .limit(4).all())
    if len(related) < 4:
        others = (Camera.query
                  .filter(Camera.brand != product.brand, Camera.type == 'Rent')
                  .limit(4 - len(related)).all())
        related += others
    return render_template('product_detail.html', product=product, related=related)


@app.route('/<slug>')
def product_alias(slug):
    cam = Camera.query.filter_by(slug=slug).first()
    if cam:
        return redirect(url_for('product_detail', slug=slug))
    return redirect(url_for('home'))


# ── Cart (online PURCHASE of for-sale cameras) ────────────────────────────────

def buy_available(camera):
    """A for-sale unit is buyable when in stock; rentals fall back to legacy stock."""
    if camera.type == 'Sale':
        return 1 if camera.sale_state == 'stock' else 0
    return camera.stock or 0


@app.route('/add_to_cart/<int:camera_id>')
def add_to_cart(camera_id):
    if 'cart' not in session:
        session['cart'] = {}
    cart = session['cart']
    camera = Camera.query.get_or_404(camera_id)
    if camera.type != 'Sale':          # rentals are booked via /rent, never purchased
        return redirect(url_for('product_detail', slug=camera.slug))
    avail = buy_available(camera)
    if avail > 0:
        key = str(camera_id)
        if cart.get(key, 0) < avail:
            cart[key] = cart.get(key, 0) + 1
            session.modified = True
    return redirect(url_for('cart'))


@app.route('/cart')
def cart():
    items, total = [], 0
    for cam_id, qty in session.get('cart', {}).items():
        camera = db.session.get(Camera, int(cam_id))
        if camera:
            item_total = camera.price * qty
            total += item_total
            items.append({'camera': camera, 'quantity': qty, 'item_total': item_total,
                          'available': buy_available(camera)})
    return render_template('cart.html', items=items, total=total)


@app.route('/checkout', methods=['POST'])
def checkout():
    cart_items = session.get('cart', {})
    if not cart_items:
        return redirect(url_for('buy_store'))
    customer_name = request.form.get('customer_name')
    phone         = request.form.get('phone')
    for cam_id in cart_items:
        camera = db.session.get(Camera, int(cam_id))
        if not camera or camera.type != 'Sale':    # only for-sale units are purchasable
            continue
        if camera.sale_state == 'stock':           # one physical unit → mark sold
            camera.sale_state = 'sold'
            camera.is_sold = True
            camera.sold_price = camera.price
            camera.sold_to = customer_name
            camera.sold_phone = phone
            camera.sold_date = datetime.now()
    db.session.commit()
    session.pop('cart', None)
    return render_template('cart.html',
                           message='Mua hàng thành công! Chúng tôi sẽ sớm liên hệ với bạn.',
                           items=[], total=0)


# ── Rental API ───────────────────────────────────────────────────────────────

@app.route('/api/camera-availability/<int:camera_id>')
def camera_availability_api(camera_id):
    """Public: only reveals whether a camera is free for a given [start,end) — not the schedule."""
    start = parse_dt(request.args.get('start'))
    end   = parse_dt(request.args.get('end'))
    if not start or not end or end <= start:
        return jsonify({'available': False, 'error': 'invalid range'}), 400
    busy = bool(bookings_overlapping(start, end, camera_id)) or \
           bool(appointments_overlapping(start, end, camera_id))
    return jsonify({'available': not busy})


@app.route('/api/camera-bookings/<int:camera_id>')
def camera_bookings_api(camera_id):
    """Public: the camera's upcoming reserved time ranges so the booking form can grey
    out taken days/times. Returns only start/end instants — no customer info."""
    now = datetime.now()
    ranges = []
    for b in RentalBooking.query.filter_by(camera_id=camera_id).filter(RentalBooking.end_date > now).all():
        if b.start_date and b.end_date:
            ranges.append({'start': b.start_date.isoformat(), 'end': b.end_date.isoformat()})
    for a in (Appointment.query.filter_by(camera_id=camera_id)
              .filter(Appointment.status != 'cancelled', Appointment.end_time > now).all()):
        if a.start_time and a.end_time:
            ranges.append({'start': a.start_time.isoformat(), 'end': a.end_time.isoformat()})
    return jsonify({'ranges': ranges})


@app.route('/api/available-cameras')
def available_cameras_api():
    start_str = request.args.get('start')
    end_str   = request.args.get('end')
    start = parse_dt(start_str)
    end   = parse_dt(end_str)
    if not start or not end:
        return jsonify({'error': 'Invalid date format'}), 400
    if end <= start:
        return jsonify({'error': 'End date must be after start date'}), 400
    busy_ids = {r[0] for r in db.session.query(RentalBooking.camera_id).filter(
        RentalBooking.start_date < end,
        RentalBooking.end_date > start,
    ).distinct().all()}
    # also treat cameras with an overlapping buy-appointment as unavailable (matches /rent)
    busy_ids |= {r[0] for r in db.session.query(Appointment.camera_id).filter(
        Appointment.camera_id.isnot(None),
        Appointment.status != 'cancelled',
        Appointment.start_time < end,
        Appointment.end_time > start,
    ).distinct().all()}
    q = Camera.query.filter_by(type='Rent').filter(Camera.is_sold.isnot(True))
    if busy_ids:
        q = q.filter(~Camera.id.in_(busy_ids))
    return jsonify([{'id': c.id, 'name': c.name, 'price': c.price} for c in q.all()])


@app.route('/rent', methods=['GET', 'POST'])
def rent():
    message = error = None
    if request.method == 'POST':
        camera = db.session.get(Camera, request.form.get('camera_id'))
        start_date = parse_dt(request.form.get('start_date'))
        end_date   = parse_dt(request.form.get('end_date'))
        if not camera or camera.type != 'Rent':
            error = 'Vui lòng chọn một máy hợp lệ để thuê.'
        elif banned_customer(request.form.get('phone')):
            error = 'Số điện thoại này hiện không thể đặt thuê online. Vui lòng liên hệ cửa hàng.'
        elif not start_date or not end_date or end_date <= start_date:
            error = 'Thời gian thuê không hợp lệ — giờ trả phải sau giờ nhận.'
        elif bookings_overlapping(start_date, end_date, camera.id) or appointments_overlapping(start_date, end_date, camera.id):
            error = f'Máy “{camera.name}” đã có người đặt trong khung giờ bạn chọn. Vui lòng chọn thời gian khác hoặc máy khác.'
        else:
            total_price = compute_rental_total(camera, start_date, end_date)
            b = RentalBooking(
                camera_id=camera.id, customer_name=(request.form.get('customer_name') or '').strip(),
                phone=(request.form.get('phone') or '').strip(), notes=(request.form.get('notes') or '').strip(),
                start_date=start_date, end_date=end_date, total_price=total_price,
            )
            db.session.add(b)
            db.session.commit()
            race_err = confirm_rental_or_rollback(b)
            if race_err:
                error = race_err
            else:
                message = f'Thành công! Bạn đã đặt thuê {camera.name}. Tổng tiền dự kiến: {total_price:,.0f}đ'.replace(',', '.')
    rent_cameras = rent_cameras_ordered()
    camera_slug = request.args.get('camera', '')
    selected_camera = Camera.query.filter_by(slug=camera_slug).first() if camera_slug else None
    # Carry over any info the homepage hero form sent via GET so the lead isn't lost.
    def _dtl(arg):                       # normalise date/datetime → datetime-local value
        d = parse_dt(request.args.get(arg))
        return d.strftime('%Y-%m-%dT%H:%M') if d else ''
    prefill = {
        'customer_name': request.args.get('customer_name', ''),
        'phone':         request.args.get('phone', ''),
        'notes':         request.args.get('notes', ''),
        'start_date':    _dtl('start_date'),
        'end_date':      _dtl('end_date'),
    }
    return render_template('rent.html', cameras=rent_cameras, message=message,
                           error=error, selected_camera=selected_camera, prefill=prefill)


# ── Admin helpers ─────────────────────────────────────────────────────────────

# Stable palette: each camera keeps the same color everywhere in the admin.
ADMIN_COLORS = [
    '#ef4444', '#3b82f6', '#10b981', '#f59e0b', '#8b5cf6',
    '#f97316', '#06b6d4', '#ec4899', '#84cc16', '#6366f1',
    '#14b8a6', '#e11d48', '#0ea5e9', '#a855f7', '#22c55e',
]

APPT_COLOR = '#0e7490'        # buy-appointments (booked): high-contrast teal
APPT_DONE_COLOR = '#16a34a'   # finished appointments turn green so they're easy to spot


def color_map_for(cameras):
    """Map camera.id → hex color, stable by sort order."""
    return {cam.id: ADMIN_COLORS[i % len(ADMIN_COLORS)] for i, cam in enumerate(cameras)}


def rent_cameras_ordered(include_sold=False):
    q = Camera.query.filter_by(type='Rent')
    if not include_sold:
        q = q.filter(Camera.is_sold.isnot(True))
    return q.order_by(Camera.brand, Camera.name).all()


def parse_dt(value):
    """Parse a datetime from several accepted shapes; date-only → 09:00."""
    if not value:
        return None
    value = value.strip()
    for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M'):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    try:
        d = datetime.strptime(value, '%Y-%m-%d')
        return d.replace(hour=9, minute=0)
    except ValueError:
        pass
    try:
        return datetime.fromisoformat(value[:19])
    except ValueError:
        return None


def compute_rental_total(camera, start, end):
    """Tiered total: day1 = price, day2 = price_d2, day3 = price_d3, day4+ = price_d4 each."""
    if not camera or not start or not end:
        return 0
    seconds = (end - start).total_seconds()
    days = max(1, math.ceil(seconds / 86400)) if seconds > 0 else 1
    total = camera.price or 0
    if days >= 2:
        total += camera.price_d2 or 0
    if days >= 3:
        total += camera.price_d3 or 0
    if days > 3:
        total += (camera.price_d4 or 0) * (days - 3)
    return int(total)


def serialize_booking(b, color=None):
    cam = b.camera
    days = max(1, math.ceil((b.end_date - b.start_date).total_seconds() / 86400))
    return {
        'id':         b.id,
        'camera_id':  b.camera_id,
        'camera':     cam.name if cam else '—',
        'brand':      cam.brand if cam else '',
        'customer':   b.customer_name or '',
        'phone':      b.phone or '',
        'notes':      b.notes or '',
        'total':      int(b.total_price or 0),
        'days':       days,
        'start':      b.start_date.isoformat(),
        'end':        b.end_date.isoformat(),
        'start_date': b.start_date.strftime('%Y-%m-%d'),
        'end_date':   b.end_date.strftime('%Y-%m-%d'),
        'color':      color or '#6b7280',
        'security_deposit': int(b.security_deposit or 0),
        'condition_out':    b.condition_out or '',
        'condition_in':     b.condition_in or '',
        'returned_at':      b.returned_at.strftime('%Y-%m-%dT%H:%M') if b.returned_at else '',
        'staff':            b.staff or '',
        'overdue':          bool(b.is_overdue),
    }


def serialize_appt(a):
    cam = a.camera
    end = a.end_time or (a.start_time + timedelta(minutes=45))
    return {
        'id':        a.id,
        'customer':  a.customer_name or '',
        'phone':     a.phone or '',
        'camera_id': a.camera_id,
        'camera':    cam.name if cam else '',
        'interest':  a.interest or '',
        'notes':     a.notes or '',
        'status':    a.status or 'booked',
        'start':     a.start_time.isoformat(),
        'end':       end.isoformat(),
        'date':      a.start_time.strftime('%Y-%m-%d'),
        'time':      a.start_time.strftime('%H:%M'),
        'kind':      'appt',
    }


def bookings_overlapping(start, end, camera_id, exclude_id=None):
    """Other bookings for the same camera that overlap [start, end)."""
    q = RentalBooking.query.filter(
        RentalBooking.camera_id == camera_id,
        RentalBooking.start_date < end,
        RentalBooking.end_date > start,
    )
    if exclude_id:
        q = q.filter(RentalBooking.id != exclude_id)
    return q.all()


def appointments_overlapping(start, end, camera_id, exclude_id=None):
    """Active appointments that reserve the SAME camera and overlap [start, end)."""
    if not camera_id:
        return []
    q = Appointment.query.filter(
        Appointment.camera_id == camera_id,
        Appointment.status != 'cancelled',
        Appointment.start_time < end,
        Appointment.end_time > start,
    )
    if exclude_id:
        q = q.filter(Appointment.id != exclude_id)
    return q.all()


def confirm_rental_or_rollback(b):
    """Race-safe finalize: after committing rental b, re-verify no *earlier* rental or any
    appointment reservation grabbed the same slot concurrently. Deterministic tiebreak — the
    smaller-id rental wins, and any unit-reserving appointment beats a rental. Returns an error
    string (and rolls b back) if b loses, else None."""
    earlier = [o for o in bookings_overlapping(b.start_date, b.end_date, b.camera_id, exclude_id=b.id) if o.id < b.id]
    appt_clash = appointments_overlapping(b.start_date, b.end_date, b.camera_id)
    if earlier or appt_clash:
        name = b.camera.name if b.camera else ''
        db.session.delete(b)
        db.session.commit()
        return f'Máy “{name}” vừa có người khác đặt trùng khung giờ. Vui lòng chọn thời gian khác.'
    return None


# ── Admin custom views (live inside the Flask-Admin panel) ─────────────────────

def banned_customer(phone):
    """The blacklisted Customer profile matching this phone, if any."""
    phone = (phone or '').strip()
    if not phone:
        return None
    return Customer.query.filter_by(phone=phone, is_banned=True).first()


class BookingGridView(StaffView):
    """Google-Sheets-style grid: rows = cameras, columns = days, cells = bookings."""

    @expose('/')
    def index(self):
        cameras   = rent_cameras_ordered()
        color_map = color_map_for(cameras)
        # All active cameras selectable when an appointment reserves a specific unit.
        appt_cameras = (Camera.query.filter(Camera.is_sold.isnot(True))
                        .order_by(Camera.type, Camera.brand, Camera.name).all())
        return self.render('admin/booking_grid.html', cameras=cameras,
                           color_map=color_map, appt_cameras=appt_cameras)

    @expose('/data')
    def data(self):
        cameras   = rent_cameras_ordered()
        color_map = color_map_for(cameras)
        start = parse_dt(request.args.get('start')) or datetime.now()
        start = start.replace(hour=0, minute=0, second=0, microsecond=0)   # align to day columns
        try:
            days = max(1, min(60, int(request.args.get('days', 14))))
        except (TypeError, ValueError):
            days = 14
        end = start + timedelta(days=days)
        bookings = (RentalBooking.query
                    .filter(RentalBooking.start_date < end, RentalBooking.end_date > start)
                    .all())
        appts = (Appointment.query
                 .filter(Appointment.status != 'cancelled',
                         Appointment.start_time >= start, Appointment.start_time < end)
                 .order_by(Appointment.start_time)
                 .all())
        return jsonify({
            'cameras':  [{'id': c.id, 'name': c.name, 'brand': c.brand,
                          'color': color_map.get(c.id, '#6b7280')} for c in cameras],
            'bookings': [serialize_booking(b, color_map.get(b.camera_id)) for b in bookings],
            'appointments': [serialize_appt(a) for a in appts],
        })

    @expose('/save', methods=['POST'])
    def save(self):
        d = request.get_json(silent=True) or {}
        cam = db.session.get(Camera,d.get('camera_id'))
        start = parse_dt(d.get('start_date'))
        end   = parse_dt(d.get('end_date'))
        if not cam:
            return jsonify({'ok': False, 'error': 'Không tìm thấy máy ảnh'}), 400
        if not start or not end or end <= start:
            return jsonify({'ok': False, 'error': 'Ngày nhận / trả không hợp lệ'}), 400

        bid = d.get('id')
        # Block double-booking the same camera for an overlapping time (exclude self on edit).
        overlaps = bookings_overlapping(start, end, cam.id, exclude_id=int(bid) if bid else None)
        if overlaps:
            names = ', '.join(o.customer_name or '?' for o in overlaps)
            return jsonify({'ok': False,
                            'error': f'Máy “{cam.name}” đã được đặt trùng khung giờ này (khách: {names}). Chọn thời gian khác.'}), 409
        # Block if a buy-appointment has reserved this exact unit for the same window.
        appt_clash = appointments_overlapping(start, end, cam.id)
        if appt_clash:
            who = ', '.join(a.customer_name or '?' for a in appt_clash)
            return jsonify({'ok': False,
                            'error': f'Máy “{cam.name}” đang được giữ cho lịch hẹn mua (khách: {who}). Chọn thời gian khác.'}), 409

        # Blacklisted renters can't book (admin can override deliberately).
        banned = banned_customer(d.get('phone'))
        if banned and not is_admin():
            return jsonify({'ok': False,
                            'error': f'Khách này đã bị chặn thuê ({banned.ban_reason or "vi phạm chính sách"}). Liên hệ quản lý.'}), 403

        if bid:
            b = db.session.get(RentalBooking,bid)
            if not b:
                return jsonify({'ok': False, 'error': 'Không tìm thấy đơn'}), 404
        else:
            b = RentalBooking(staff=current_staff_user())
            db.session.add(b)

        b.camera_id     = cam.id
        b.customer_name = (d.get('customer_name') or '').strip()
        b.phone         = (d.get('phone') or '').strip()
        b.notes         = (d.get('notes') or '').strip()
        b.start_date    = start
        b.end_date      = end
        b.total_price   = compute_rental_total(cam, start, end)
        # rental-agreement operations: security deposit + equipment condition
        if 'security_deposit' in d:
            try:
                b.security_deposit = max(0, int(d.get('security_deposit') or 0))
            except (TypeError, ValueError):
                pass
        if 'condition_out' in d:
            b.condition_out = (d.get('condition_out') or '').strip()[:300]
        if 'condition_in' in d:
            b.condition_in = (d.get('condition_in') or '').strip()[:300]
        db.session.commit()

        race_err = confirm_rental_or_rollback(b)
        if race_err:
            return jsonify({'ok': False, 'error': race_err}), 409

        cmap = color_map_for(rent_cameras_ordered())
        return jsonify({'ok': True, 'booking': serialize_booking(b, cmap.get(cam.id))})

    @expose('/delete', methods=['POST'])
    def delete(self):
        d = request.get_json(silent=True) or {}
        b = db.session.get(RentalBooking,d.get('id'))
        if not b:
            return jsonify({'ok': False, 'error': 'Không tìm thấy đơn'}), 404
        db.session.delete(b)
        db.session.commit()
        return jsonify({'ok': True})

    @expose('/return', methods=['POST'])
    def mark_return(self):
        """Check a rental back in: record the equipment's condition and the return
        time (POST returned=false to undo an accidental check-in)."""
        d = request.get_json(silent=True) or {}
        b = db.session.get(RentalBooking, d.get('id'))
        if not b:
            return jsonify({'ok': False, 'error': 'Không tìm thấy đơn'}), 404
        if d.get('returned') is False:
            b.returned_at = None
        else:
            b.returned_at = parse_dt(d.get('returned_at')) or datetime.now()
            if 'condition_in' in d:
                b.condition_in = (d.get('condition_in') or '').strip()[:300]
        db.session.commit()
        cmap = color_map_for(rent_cameras_ordered())
        return jsonify({'ok': True, 'booking': serialize_booking(b, cmap.get(b.camera_id))})

    # ── Combined calendar feed: rentals + buy-appointments ───────────────────
    @expose('/events')
    def events(self):
        show = request.args.get('show', 'both')
        cameras   = rent_cameras_ordered()
        color_map = color_map_for(cameras)
        start = parse_dt((request.args.get('start') or '')[:19])
        end   = parse_dt((request.args.get('end') or '')[:19])
        events = []
        if show in ('both', 'rent'):
            q = RentalBooking.query
            if start: q = q.filter(RentalBooking.end_date >= start)
            if end:   q = q.filter(RentalBooking.start_date <= end)
            for b in q.all():
                color = color_map.get(b.camera_id, '#6b7280')
                cam = b.camera
                events.append({
                    'id': f'r{b.id}',
                    'title': f'{cam.name if cam else "Máy?"} · {b.customer_name}',
                    'start': b.start_date.isoformat(), 'end': b.end_date.isoformat(),
                    'backgroundColor': color, 'borderColor': color,
                    'extendedProps': serialize_booking(b, color) | {'kind': 'rent'},
                })
        if show in ('both', 'appt'):
            qa = Appointment.query.filter(Appointment.status != 'cancelled')
            if start: qa = qa.filter(Appointment.start_time >= start - timedelta(days=1))
            if end:   qa = qa.filter(Appointment.start_time <= end)
            for a in qa.all():
                done = (a.status == 'done')
                color = APPT_DONE_COLOR if done else APPT_COLOR
                title = ('✅ ' if done else '🛒 ') + (a.customer_name or 'Khách') + (f' · {a.interest}' if a.interest else '')
                events.append({
                    'id': f'a{a.id}', 'title': title,
                    'start': a.start_time.isoformat(),
                    'end': (a.end_time or (a.start_time + timedelta(minutes=45))).isoformat(),
                    'backgroundColor': color, 'borderColor': color,
                    'extendedProps': serialize_appt(a),
                })
        return jsonify(events)

    # ── Buy-appointment CRUD ─────────────────────────────────────────────────
    @expose('/appt/save', methods=['POST'])
    def appt_save(self):
        d = request.get_json(silent=True) or {}
        start = parse_dt(d.get('start_time'))
        if not start:
            return jsonify({'ok': False, 'error': 'Thời gian hẹn không hợp lệ'}), 400
        aid = d.get('id')
        end = parse_dt(d.get('end_time'))
        end = end if (end and end > start) else (start + timedelta(minutes=45))

        # Optional: reserve a specific unit. If set, block if that unit is rented or
        # already reserved by another appointment for an overlapping window.
        cam_id = d.get('camera_id') or None
        status = d.get('status') if d.get('status') in ('booked', 'done', 'cancelled') else 'booked'
        if cam_id and status != 'cancelled':
            cam = db.session.get(Camera, cam_id)
            if not cam:
                return jsonify({'ok': False, 'error': 'Không tìm thấy máy'}), 400
            if bookings_overlapping(start, end, cam.id):
                return jsonify({'ok': False, 'error': f'Máy “{cam.name}” đang có đơn thuê trùng khung giờ này.'}), 409
            appt_clash = appointments_overlapping(start, end, cam.id, exclude_id=int(aid) if aid else None)
            if appt_clash:
                who = ', '.join(x.customer_name or '?' for x in appt_clash)
                return jsonify({'ok': False, 'error': f'Máy “{cam.name}” đã được giữ cho lịch hẹn khác (khách: {who}).'}), 409

        if aid:
            a = db.session.get(Appointment, aid)
            if not a:
                return jsonify({'ok': False, 'error': 'Không tìm thấy lịch hẹn'}), 404
        else:
            a = Appointment()
            db.session.add(a)
        a.customer_name = (d.get('customer_name') or '').strip()
        a.phone         = (d.get('phone') or '').strip()
        a.interest      = (d.get('interest') or '').strip()
        a.notes         = (d.get('notes') or '').strip()
        a.start_time    = start
        a.end_time      = end
        a.camera_id     = int(cam_id) if cam_id else None
        a.status        = status
        db.session.commit()
        return jsonify({'ok': True, 'appt': serialize_appt(a)})

    @expose('/appt/delete', methods=['POST'])
    def appt_delete(self):
        d = request.get_json(silent=True) or {}
        a = db.session.get(Appointment, d.get('id'))
        if not a:
            return jsonify({'ok': False, 'error': 'Không tìm thấy lịch hẹn'}), 404
        db.session.delete(a)
        db.session.commit()
        return jsonify({'ok': True})


class CalendarView(StaffView):
    """FullCalendar week / 3-day time-grid view."""

    @expose('/')
    def index(self):
        cameras   = rent_cameras_ordered()
        color_map = color_map_for(cameras)
        return self.render('admin/calendar.html', cameras=cameras, color_map=color_map)

    @expose('/events')
    def events(self):
        cameras   = rent_cameras_ordered()
        color_map = color_map_for(cameras)
        start = parse_dt((request.args.get('start') or '')[:19])
        end   = parse_dt((request.args.get('end') or '')[:19])
        q = RentalBooking.query
        if start:
            q = q.filter(RentalBooking.end_date >= start)
        if end:
            q = q.filter(RentalBooking.start_date <= end)
        events = []
        for b in q.all():
            color = color_map.get(b.camera_id, '#6b7280')
            cam   = b.camera
            events.append({
                'id':              b.id,
                'title':           f'{cam.name if cam else "Máy?"} · {b.customer_name}',
                'start':           b.start_date.isoformat(),
                'end':             b.end_date.isoformat(),
                'backgroundColor': color,
                'borderColor':     color,
                'extendedProps':   serialize_booking(b, color),
            })
        return jsonify(events)


def serialize_unit(cam):
    """A single for-sale camera unit (Selling ledger row)."""
    return {
        'id': cam.id, 'name': cam.name, 'brand': cam.brand or '',
        'origin': cam.origin or '', 'accessory': cam.accessory or '', 'color': cam.color or '',
        'category': cam.category or '',
        'date_in': cam.date_in.strftime('%Y-%m-%d') if cam.date_in else '',
        'note': cam.description or '', 'sold_price': cam.sold_price or 0,
        'deposit_amount': cam.deposit_amount or 0,
        'import_cost': cam.import_cost or 0, 'repair_cost': cam.repair_cost or 0, 'profit': cam.profit,
        'sale_state': cam.sale_state or 'stock',
        'sold_to': cam.sold_to or '', 'sold_phone': cam.sold_phone or '', 'sold_source': cam.sold_source or '',
        'sold_note': cam.sold_note or '',
        'sold_at': cam.sold_date.strftime('%Y-%m-%dT%H:%M') if cam.sold_date else '',
        'sold_date': cam.sold_date.strftime('%Y-%m-%d') if cam.sold_date else '',
        'sold_month': cam.sold_date.strftime('%Y-%m') if cam.sold_date else '',
        'state_label': cam.state_label, 'sold_by': cam.sold_by or '',
        'gift_cost': cam.gift_cost, 'gift_label': cam.gift_label, 'gifts': cam.gifts,
    }


# ── Financial redaction: employees never see acquisition costs or profit ──────
def redact_unit(u):
    """Employee-safe copy of a serialized camera unit."""
    if is_admin():
        return u
    u = dict(u)
    for k in ('import_cost', 'repair_cost', 'profit', 'gift_cost'):
        u.pop(k, None)
    u['gifts'] = [{'id': g.get('id'), 'name': g.get('name')} for g in (u.get('gifts') or [])]
    return u


def redact_pnl(p):
    """Employee-safe copy of a camera P&L dict."""
    if is_admin():
        return p
    p = dict(p)
    for k in ('cost', 'profit', 'inventory_value'):
        p.pop(k, None)
    return p


def redact_accessory(a):
    """Employee-safe copy of a serialized accessory (no cost / stock value)."""
    if is_admin():
        return a
    a = dict(a)
    for k in ('cost', 'stock_value'):
        a.pop(k, None)
    return a


def redact_accessory_sale(s):
    """Employee-safe copy of a logged accessory sale (no COGS / profit)."""
    if is_admin():
        return s
    s = dict(s)
    for k in ('unit_cost', 'profit'):
        s.pop(k, None)
    return s


def redact_accessory_damage(x):
    """Employee-safe copy of a damaged-accessory record (no cost / loss figures)."""
    if is_admin():
        return x
    x = dict(x)
    for k in ('unit_cost', 'loss'):
        x.pop(k, None)
    return x


def serialize_cost(c):
    return {'id': c.id, 'category': c.category or 'Khác', 'note': c.note or '',
            'amount': c.amount or 0,
            'date': c.cost_date.strftime('%Y-%m-%d') if c.cost_date else ''}


def serialize_accessory(a):
    return {'id': a.id, 'name': a.name, 'category': a.category or '',
            'cost': a.cost or 0, 'price': a.price or 0, 'stock': a.stock or 0,
            'note': a.note or '', 'stock_value': a.stock_value}


def serialize_accessory_sale(s):
    return {'id': s.id, 'name': s.name or '', 'quantity': s.quantity or 0,
            'unit_price': s.unit_price or 0, 'unit_cost': s.unit_cost or 0,
            'revenue': s.revenue, 'profit': s.profit, 'note': s.note or '',
            'customer_name': s.customer_name or '', 'phone': s.phone or '',
            'date': s.sale_date.strftime('%d/%m/%Y %H:%M') if s.sale_date else '',
            'date_iso': s.sale_date.strftime('%Y-%m-%d') if s.sale_date else ''}


def serialize_accessory_damage(x):
    return {'id': x.id, 'name': x.name or '', 'quantity': x.quantity or 0,
            'unit_cost': x.unit_cost or 0, 'loss': x.loss, 'note': x.note or '',
            'staff': x.staff or '',
            'date': x.damage_date.strftime('%d/%m/%Y %H:%M') if x.damage_date else '',
            'date_iso': x.damage_date.strftime('%Y-%m-%d') if x.damage_date else ''}


def camera_pnl(cam):
    return {
        'units_sold': cam.units_sold, 'rental_revenue': cam.rental_revenue,
        'sale_revenue': cam.sale_revenue, 'revenue': cam.revenue,
        'cost': cam.cost, 'profit': cam.profit, 'inventory_value': cam.inventory_value,
        'sale_state': cam.sale_state, 'state_label': cam.state_label,
    }


class CamerasView(StaffView):
    """Inventory manager. Renting tab = editable table; Selling tab = per-color/status stock.
    Employees get read-only inventory + POS selling; pricing, costs and stock
    overrides stay admin-only."""

    # Whitelist of editable fields → caster (prevents mass-assignment).
    EDITABLE = {
        'name':        str,   'brand':       str,   'badge':       str,   'type': str,
        'category':    str,   'description': str,   'color':       str,
        'origin':      str,   'accessory':   str,   'sold_to':     str,   'sold_phone': str,
        'sold_source': str,
        'price':       int,   'price_d2':    int,   'price_d3':    int,  'price_d4': int,
        'stock':       int,   'import_cost': int,   'repair_cost': int,  'sold_price': int,
        'deposit_amount': int, 'reorder_point': int,
        'is_broken':   bool,  'is_sold':     bool,  'featured':    bool,
    }
    # Operational, non-financial fields an employee may still inline-edit
    # (customer info on a sale they processed, cosmetic details). Pricing, costs
    # and stock overrides are admin-only.
    EMPLOYEE_EDITABLE = ('sold_to', 'sold_phone', 'sold_source',
                         'description', 'accessory', 'color')
    # Rental pricing an employee may adjust — but only on rental (Rent) units.
    # Sale pricing (sold_price) still goes through the guarded /sell flow.
    EMPLOYEE_RENT_PRICE = ('price', 'price_d2', 'price_d3', 'price_d4')
    SALE_STATES = ('stock', 'processing', 'deposit', 'sold', 'fixing', 'unfixable')
    SELL_STATES = ('sold', 'deposit')    # states that finalise a sale (capture price/customer)

    @expose('/')
    def index(self):
        rent = Camera.query.filter_by(type='Rent').order_by(Camera.brand, Camera.name).all()
        # Active selling inventory = stock / processing / fixing / unfixable (not sold, not deposited).
        sale = (Camera.query.filter_by(type='Sale')
                .filter(~Camera.sale_state.in_(('sold', 'deposit'))).order_by(Camera.id).all())
        # Sold units live in their own tab, newest sale first.
        sold = (Camera.query.filter_by(type='Sale', sale_state='sold')
                .order_by(Camera.sold_date.desc(), Camera.id.desc()).all())
        # Deposited (đã cọc) units — reserved for online buyers, revenue already booked.
        deposit = (Camera.query.filter_by(type='Sale', sale_state='deposit')
                   .order_by(Camera.sold_date.desc(), Camera.id.desc()).all())
        color_map = color_map_for(rent_cameras_ordered())   # rental colors match the grid
        brands = sorted({c.brand for c in (rent + sale + sold + deposit) if c.brand})
        gift_options = (Accessory.query.filter(Accessory.stock > 0)
                        .order_by(Accessory.category, Accessory.name).all())
        return self.render('admin/cameras.html',
                           rent_cameras=rent, sale_cameras=sale, sold_cameras=sold,
                           deposit_cameras=deposit,
                           cameras=rent + sale + sold + deposit, color_map=color_map,
                           categories=CAMERA_CATEGORIES, brands=brands,
                           sale_sources=SALE_SOURCES,
                           today=datetime.now().strftime('%Y-%m-%d'),
                           this_month=datetime.now().strftime('%Y-%m'),
                           this_year=datetime.now().strftime('%Y'),
                           gift_options=[redact_accessory(serialize_accessory(a)) for a in gift_options])

    @expose('/update', methods=['POST'])
    def update(self):
        d = request.get_json(silent=True) or {}
        cam = db.session.get(Camera, d.get('id'))
        field, value = d.get('field'), d.get('value')
        if not cam:
            return jsonify({'ok': False, 'error': 'Không tìm thấy máy'}), 404

        # Employees: read-only inventory except a few operational fields.
        # sale_state is allowed here (POS + re-classify); the one restriction —
        # un-finalising a sale — is enforced inside the sale_state branch below.
        # Rental prices are also editable by employees, but only on Rent units.
        if not is_admin() and field != 'sale_state' and field not in self.EMPLOYEE_EDITABLE:
            allow_rent_price = (field in self.EMPLOYEE_RENT_PRICE and cam.type == 'Rent')
            if not allow_rent_price:
                return _forbidden('Chỉ quản lý được sửa trường này.')

        # ── per-unit special fields ──
        if field == 'date_in':
            cam.date_in = parse_dt(value) if value else None
            db.session.commit()
            return jsonify({'ok': True, 'pnl': redact_pnl(camera_pnl(cam))})
        if field == 'sold_date':
            cam.sold_date = parse_dt(value) if value else None
            db.session.commit()
            return jsonify({'ok': True, 'pnl': redact_pnl(camera_pnl(cam))})
        if field == 'sale_state':
            if value not in self.SALE_STATES:
                return jsonify({'ok': False, 'error': 'Trạng thái không hợp lệ'}), 400
            # Employees may sell and re-classify freely, but must NOT un-finalise a
            # sale: đã bán / đã cọc → còn hàng / đang xử lý is admin-only.
            if (not is_admin() and cam.sale_state in Camera.SOLD_STATES
                    and value in ('stock', 'processing')):
                return _forbidden('Chỉ quản lý được chuyển máy “đã bán / đã cọc” '
                                  'về “còn hàng / đang xử lý”.')
            # A sale must have a price: block finalising (đã bán / đã cọc) a 0đ unit.
            if value in self.SELL_STATES and (cam.sold_price or 0) <= 0:
                return jsonify({'ok': False,
                                'error': 'Cần nhập giá bán (khác 0đ) trước khi đánh dấu '
                                         '“đã bán / đã cọc”. Dùng nút “Bán” để nhập giá.'}), 400
            cam.sale_state = value
            cam.is_sold = (value == 'sold')
            if value != 'deposit':                 # deposit amount only applies while đã cọc
                cam.deposit_amount = 0
            if value in self.SELL_STATES:          # sold / đã cọc → realised sale
                if not cam.sold_date:
                    cam.sold_date = datetime.now()
            else:                                  # reverting out of a sale
                cam.sold_date = None
                if value == 'stock':
                    # fully undo the sale: restock gifts, drop their history, clear sale info
                    self._restock_gifts(cam)
                    cam.gift_json = ''
                    cam.sold_price = 0
                    cam.sold_to = cam.sold_phone = cam.sold_source = cam.sold_note = ''
            db.session.commit()
            return jsonify({'ok': True, 'pnl': redact_pnl(camera_pnl(cam)), 'unit': redact_unit(serialize_unit(cam))})

        if field not in self.EDITABLE:
            return jsonify({'ok': False, 'error': 'Trường không hợp lệ'}), 400
        caster = self.EDITABLE[field]
        try:
            if caster is bool:
                cast = bool(value) if isinstance(value, bool) else str(value).lower() in ('1', 'true', 'on', 'yes')
            elif caster is int:
                cast = int(value or 0)
            else:
                cast = (str(value) if value is not None else '').strip()
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'Giá trị không hợp lệ'}), 400
        if field == 'type' and cast not in ('Rent', 'Sale'):
            return jsonify({'ok': False, 'error': 'Loại máy không hợp lệ'}), 400
        # Rental repair cost is cumulative & undated on the camera; log each change as a
        # dated expense so it lands in the Finance period it was incurred.
        if field == 'repair_cost' and cam.type == 'Rent':
            delta = int(cast) - int(cam.repair_cost or 0)
            if delta != 0:
                db.session.add(RepairLog(camera_id=cam.id, amount=delta,
                                         note=cam.name or '', created_at=datetime.now()))
        setattr(cam, field, cast)
        if field == 'is_sold':
            cam.sold_date = datetime.now() if cast else None
        db.session.commit()
        return jsonify({'ok': True, 'pnl': redact_pnl(camera_pnl(cam))})

    # ── Gift helpers (shared by sell + gift editing + revert) ────────────────
    @staticmethod
    def _restock_gifts(cam):
        """Put every gifted accessory back into stock and delete its history log."""
        for g in cam.gifts:
            acc = db.session.get(Accessory, g.get('id'))
            if acc:
                acc.stock = (acc.stock or 0) + 1
            sid = g.get('sale_id')
            if sid:
                rec = db.session.get(AccessorySale, sid)
                if rec:
                    db.session.delete(rec)

    @staticmethod
    def _apply_gifts(cam, gift_ids):
        """Deduct the chosen accessories from stock and log each as a (0đ) accessory sale.
        Price/cost = 0 because the gift cost is already booked against the camera's profit
        via gift_cost — logging it again at cost would double-count it in Finance.
        Returns the gift_json list. Assumes any previous gifts were already restocked."""
        gifts = []
        for aid in (gift_ids or []):
            acc = db.session.get(Accessory, aid)
            if acc and (acc.stock or 0) > 0:
                acc.stock -= 1
                rec = AccessorySale(accessory_id=acc.id, name=acc.name, quantity=1,
                                    unit_price=0, unit_cost=0,
                                    note=f'🎁 Tặng kèm máy: {cam.name}', sale_date=datetime.now())
                db.session.add(rec)
                db.session.flush()   # need rec.id to unlink it later on return/edit
                gifts.append({'id': acc.id, 'name': acc.name,
                              'cost': int(acc.cost or 0), 'sale_id': rec.id})
        return gifts

    @expose('/sell', methods=['POST'])
    def sell(self):
        """Finalise a for-sale unit: sold (đã bán) or deposit (đã cọc), capturing
        price + customer info. A deposit books the FULL price into that day's finance;
        deposit_amount is recorded separately (how much cọc was collected) for reference."""
        d = request.get_json(silent=True) or {}
        cam = db.session.get(Camera, d.get('id'))
        if not cam:
            return jsonify({'ok': False, 'error': 'Không tìm thấy máy'}), 404
        try:
            price = max(0, int(d.get('sold_price') or 0))
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'Giá bán không hợp lệ'}), 400
        if price <= 0:
            return jsonify({'ok': False, 'error': 'Giá bán phải lớn hơn 0đ.'}), 400
        try:
            deposit = max(0, int(d.get('deposit_amount') or 0))
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'Số tiền cọc không hợp lệ'}), 400
        # Employees may only apply the standard, pre-approved discount: the sale
        # price can't drop below listed price minus max_discount_pct (admin-set).
        if not is_admin() and (cam.price or 0) > 0:
            pct = setting_float('max_discount_pct')
            floor_price = int((cam.price or 0) * (100 - pct) / 100)
            if price < floor_price:
                return _forbidden(f'Giá bán thấp hơn mức cho phép (giảm tối đa {pct:g}% '
                                  f'→ tối thiểu {floor_price:,.0f}đ). Liên hệ quản lý.'.replace(',', '.'))
        state = d.get('state') if d.get('state') in self.SELL_STATES else 'sold'
        cam.sale_state = state
        cam.is_sold = (state == 'sold')
        cam.sold_price = price
        # deposit amount only meaningful while state='deposit'; a direct sale clears it
        cam.deposit_amount = deposit if state == 'deposit' else 0
        cam.sold_to = (d.get('customer_name') or '').strip()
        cam.sold_phone = (d.get('phone') or '').strip()
        cam.sold_source = (d.get('source') or '').strip()
        cam.sold_note = (d.get('note') or '').strip()
        cam.sold_by = current_staff_user()
        cam.sold_date = datetime.now()
        # gifted accessories: restock any previous gifts first, then apply the new selection
        self._restock_gifts(cam)
        gifts = self._apply_gifts(cam, d.get('gifts'))
        cam.gift_json = json.dumps(gifts, ensure_ascii=False) if gifts else ''
        db.session.commit()
        return jsonify({'ok': True, 'pnl': redact_pnl(camera_pnl(cam)), 'unit': redact_unit(serialize_unit(cam))})

    @expose('/gifts', methods=['POST'])
    def gifts_update(self):
        """Edit the gifted accessories on an already-sold/deposited unit."""
        d = request.get_json(silent=True) or {}
        cam = db.session.get(Camera, d.get('id'))
        if not cam:
            return jsonify({'ok': False, 'error': 'Không tìm thấy máy'}), 404
        self._restock_gifts(cam)                          # undo old gifts
        gifts = self._apply_gifts(cam, d.get('gifts'))    # apply new selection
        cam.gift_json = json.dumps(gifts, ensure_ascii=False) if gifts else ''
        db.session.commit()
        return jsonify({'ok': True, 'pnl': redact_pnl(camera_pnl(cam)), 'unit': redact_unit(serialize_unit(cam))})

    @expose('/add', methods=['POST'])
    def add(self):
        d = request.get_json(silent=True) or {}
        cam_type = d.get('type') if d.get('type') in ('Rent', 'Sale') else 'Rent'
        # Employees may add rental machines only, and never set acquisition cost
        # (a financial field they don't see). Full add (incl. sale units) is admin-only.
        if not is_admin():
            if cam_type != 'Rent':
                return _forbidden('Nhân viên chỉ được thêm máy cho thuê.')
        name = (d.get('name') or '').strip()
        if not name:
            return jsonify({'ok': False, 'error': 'Tên máy bắt buộc'}), 400
        base = ''.join(ch.lower() if ch.isalnum() else '-' for ch in name).strip('-')
        slug, n = base or 'may', 1
        while Camera.query.filter_by(slug=slug).first():
            n += 1
            slug = f'{base}-{n}'
        category = (d.get('category') or '').strip()
        state = d.get('sale_state') if d.get('sale_state') in self.SALE_STATES else 'stock'
        import_cost = int(d.get('import_cost') or 0) if is_admin() else 0
        cam = Camera(name=name, slug=slug, brand=(d.get('brand') or '').strip(),
                     type=cam_type, category=category, price=int(d.get('price') or 0),
                     import_cost=import_cost, stock=1,
                     origin=(d.get('origin') or '').strip(),
                     accessory=(d.get('accessory') or '').strip(),
                     color=(d.get('color') or '').strip(),
                     description=(d.get('note') or '').strip(),
                     sale_state=state)
        if d.get('date_in'):
            cam.date_in = parse_dt(d.get('date_in'))
        elif cam_type == 'Sale':
            cam.date_in = datetime.now()
        db.session.add(cam)
        db.session.commit()
        return jsonify({'ok': True, 'id': cam.id, 'camera': serialize_unit(cam)})

    # ── Excel import (bulk) + downloadable template ──────────────────────────
    @expose('/import-template')
    def import_template(self):
        """Download a blank .xlsx the shop can fill in and re-upload."""
        if not is_admin():
            return _forbidden('Chỉ quản lý được nhập kho hàng loạt.')
        if not HAS_OPENPYXL:
            return jsonify({'ok': False, 'error': 'Thiếu thư viện openpyxl trên máy chủ.'}), 500
        wb = Workbook()
        ws = wb.active
        ws.title = 'Nhập máy'
        head_fill = PatternFill('solid', fgColor='B91C1C')
        head_font = Font(bold=True, color='FFFFFF')
        for ci, title in enumerate(IMPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=ci, value=title)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[cell.column_letter].width = max(14, len(title) + 4)
        # one example row so the format is obvious — name flags it as a sample to delete,
        # so it's never silently imported as a real camera if the operator forgets.
        ws.append(['VÍ DỤ (xoá dòng này) LUMIX FX60', 'Bạc', 'Compact', 'N', 'Flash, pin',
                   '2026-06-01', 'Máy đẹp', 1750000, '', 'Còn hàng'])
        ws.freeze_panes = 'A2'
        # "Trạng thái" (last column) → in-cell dropdown so the shop picks a valid value
        state_col = ws.cell(row=1, column=len(IMPORT_COLUMNS)).column_letter
        dv = DataValidation(
            type='list',
            formula1='"{}"'.format(','.join(STATE_CHOICES)),
            allow_blank=True,
            showDropDown=False,           # False = show the dropdown arrow (openpyxl quirk)
        )
        dv.error = 'Chọn một trạng thái từ danh sách.'
        dv.errorTitle = 'Trạng thái không hợp lệ'
        dv.prompt = 'Chọn trạng thái từ danh sách'
        dv.promptTitle = 'Trạng thái'
        ws.add_data_validation(dv)
        dv.add('{col}2:{col}1000'.format(col=state_col))   # apply to the data rows
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='mau-nhap-may.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @expose('/import-excel', methods=['POST'])
    def import_excel(self):
        """Bulk-create for-sale units from an uploaded .xlsx (same columns as the template)."""
        if not is_admin():
            return _forbidden('Chỉ quản lý được nhập kho hàng loạt.')
        if not HAS_OPENPYXL:
            return jsonify({'ok': False, 'error': 'Thiếu thư viện openpyxl trên máy chủ.'}), 500
        f = request.files.get('file')
        if not f or not f.filename:
            return jsonify({'ok': False, 'error': 'Chưa chọn file.'}), 400
        try:
            wb = load_workbook(io.BytesIO(f.read()), data_only=True)
        except Exception:
            return jsonify({'ok': False, 'error': 'Không đọc được file Excel (.xlsx).'}), 400
        ws = wb.active
        added, errors = 0, []
        rows = list(ws.iter_rows(values_only=True))
        for ri, row in enumerate(rows, start=1):
            if ri == 1:
                continue                                   # header
            cells = list(row) + [None] * (len(IMPORT_COLUMNS) - len(row))
            name = (str(cells[0]).strip() if cells[0] is not None else '')
            if not name:
                continue                                   # skip blank lines
            try:
                base = ''.join(ch.lower() if ch.isalnum() else '-' for ch in name).strip('-')
                slug, n = base or 'may', 1
                while Camera.query.filter_by(slug=slug).first():
                    n += 1
                    slug = f'{base}-{n}'
                state = normalize_state(cells[9])
                sold_price = _to_int(cells[8])
                date_in = _cell_date(cells[5]) or datetime.now()
                cam = Camera(
                    name=name, slug=slug, type='Sale',
                    color=(str(cells[1]).strip() if cells[1] else ''),
                    category=(str(cells[2]).strip() if cells[2] else ''),
                    origin=normalize_origin(cells[3]),
                    accessory=(str(cells[4]).strip() if cells[4] else ''),
                    date_in=date_in,
                    description=(str(cells[6]).strip() if cells[6] else ''),
                    import_cost=_to_int(cells[7]),
                    sold_price=sold_price,
                    sale_state=state, price=sold_price or 0, stock=1)
                cam.is_sold = (state == 'sold')
                if state in ('sold', 'deposit'):   # realised sale → stamp the date revenue is counted
                    cam.sold_date = date_in
                db.session.add(cam)
                added += 1
            except Exception as e:                         # noqa: BLE001
                errors.append(f'Dòng {ri}: {e}')
        db.session.commit()
        return jsonify({'ok': True, 'added': added, 'errors': errors[:10]})

    @expose('/delete', methods=['POST'])
    def delete(self):
        if not is_admin():
            return _forbidden('Chỉ quản lý được xoá máy khỏi kho.')
        d = request.get_json(silent=True) or {}
        cam = db.session.get(Camera, d.get('id'))
        if not cam:
            return jsonify({'ok': False, 'error': 'Không tìm thấy máy'}), 404
        self._restock_gifts(cam)     # return any gifted accessories to stock (same as bulk_delete)
        RentalBooking.query.filter_by(camera_id=cam.id).delete()
        db.session.delete(cam)   # variants cascade via relationship
        db.session.commit()
        return jsonify({'ok': True})

    @expose('/bulk-delete', methods=['POST'])
    def bulk_delete(self):
        """Delete several cameras at once (multi-select on the Selling / Sold tabs)."""
        if not is_admin():
            return _forbidden('Chỉ quản lý được xoá máy khỏi kho.')
        d = request.get_json(silent=True) or {}
        ids = d.get('ids') or []
        try:
            ids = [int(x) for x in ids]
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'Danh sách không hợp lệ'}), 400
        deleted = 0
        for cid in ids:
            cam = db.session.get(Camera, cid)
            if not cam:
                continue
            self._restock_gifts(cam)     # return any gifted accessories to stock
            RentalBooking.query.filter_by(camera_id=cam.id).delete()
            db.session.delete(cam)       # variants cascade via relationship
            deleted += 1
        db.session.commit()
        return jsonify({'ok': True, 'deleted': deleted})

class FinanceView(AdminOnlyView):
    """Revenue / cost / profit dashboard (admin only)."""

    @staticmethod
    def _bucket(period, dt):
        """[start, end) of the period that `dt` falls into."""
        if period == 'day':
            s = dt.replace(hour=0, minute=0, second=0, microsecond=0)
            return s, s + timedelta(days=1)
        if period == 'year':
            return datetime(dt.year, 1, 1), datetime(dt.year + 1, 1, 1)
        end_y, end_m = (dt.year + 1, 1) if dt.month == 12 else (dt.year, dt.month + 1)
        return datetime(dt.year, dt.month, 1), datetime(end_y, end_m, 1)

    @staticmethod
    def _shift(period, dt, n):
        """Move the anchor date back `n` whole periods."""
        if period == 'day':
            return dt - timedelta(days=n)
        if period == 'year':
            return dt.replace(year=dt.year - n)
        total = (dt.year * 12 + (dt.month - 1)) - n
        return datetime(total // 12, total % 12 + 1, 1)

    @expose('/')
    def index(self):
        period = request.args.get('period', 'month')
        if period not in ('day', 'month', 'year', 'all'):
            period = 'month'
        now = datetime.now()

        # ── anchor: the specific day / month / year being viewed ──
        anchor_raw = (request.args.get('anchor') or '').strip()
        anchor = parse_dt(anchor_raw)
        if not anchor:
            if period == 'year' and anchor_raw.isdigit():
                anchor = datetime(int(anchor_raw), 1, 1)
            elif period == 'month' and len(anchor_raw) >= 7 and anchor_raw[:4].isdigit():
                try:
                    anchor = datetime(int(anchor_raw[:4]), int(anchor_raw[5:7]), 1)
                except ValueError:
                    anchor = None
        if not anchor:
            anchor = now
        if period == 'all':
            astart, aend = datetime(1970, 1, 1), datetime(2999, 1, 1)   # everything
        else:
            astart, aend = self._bucket(period, anchor)

        cameras = Camera.query.order_by(Camera.brand, Camera.name).all()
        rent_cameras = [c for c in cameras if c.type == 'Rent']
        sale_cameras = [c for c in cameras if c.type == 'Sale']
        bookings  = RentalBooking.query.all()
        # 'deposit' (đã cọc) counts as realised revenue too — customer has paid.
        sold_cams = [c for c in sale_cameras if c.sale_state in ('sold', 'deposit') and c.sold_date]
        store_costs = StoreCost.query.all()
        acc_sales = AccessorySale.query.all()
        acc_damages = AccessoryDamage.query.all()     # dated accessory write-offs (phụ kiện hư)
        repair_logs = RepairLog.query.all()          # dated rental-repair expenses

        def window(start, end):
            """All money flows whose own date falls inside [start, end)."""
            rrev  = sum(int(b.total_price or 0) for b in bookings if b.start_date and start <= b.start_date < end)
            rcost = sum(int(l.amount or 0) for l in repair_logs if l.created_at and start <= l.created_at < end)
            sold  = [c for c in sold_cams if start <= c.sold_date < end]
            srev  = sum(int(c.sale_revenue) for c in sold)     # deposit books full price on cọc day
            scogs = sum(int(c.cost) for c in sold)
            scost = sum(int(x.amount or 0) for x in store_costs if x.cost_date and start <= x.cost_date < end)
            asales = [s for s in acc_sales if s.sale_date and start <= s.sale_date < end]
            arev  = sum(s.revenue for s in asales)
            acogs = sum(s.cost_total for s in asales)
            # Damaged accessories: a write-off loss (cost of the units binned), booked
            # in the period they were recorded. Reduces accessory profit like a COGS.
            aloss = sum(x.loss for x in acc_damages if x.damage_date and start <= x.damage_date < end)
            return {'rrev': rrev, 'rcost': rcost, 'srev': srev, 'scogs': scogs, 'sprof': srev - scogs,
                    'scost': scost, 'sold': sold,
                    'arev': arev, 'acogs': acogs, 'aloss': aloss, 'aprof': arev - acogs - aloss}

        # ── trailing buckets for the trend charts (ending at the anchor) ──
        # 'all' has no single anchor → show a yearly trend over the last few years for context.
        chart_period = 'year' if period == 'all' else period
        chart_anchor = now if period == 'all' else anchor
        counts = {'day': 14, 'month': 12, 'year': 5}[chart_period]
        buckets = []
        for i in range(counts - 1, -1, -1):
            bdt = self._shift(chart_period, chart_anchor, i)
            bs, be = self._bucket(chart_period, bdt)
            label = bs.strftime('%d/%m') if chart_period == 'day' else (str(bs.year) if chart_period == 'year' else f'{bs.month:02d}/{bs.year}')
            buckets.append((label, bs, be, period != 'all' and bs == astart))
        rental_series, sale_rev_series, sale_profit_series, store_cost_series = [], [], [], []
        for (label, bs, be, sel) in buckets:
            w = window(bs, be)
            rental_series.append({'label': label, 'value': int(w['rrev']), 'sel': sel})
            sale_rev_series.append({'label': label, 'value': int(w['srev']), 'sel': sel})
            sale_profit_series.append({'label': label, 'value': int(w['sprof']), 'sel': sel})
            store_cost_series.append({'label': label, 'value': int(w['scost']), 'sel': sel})

        # ── totals for the SELECTED period only (dates now matter) ──
        w = window(astart, aend)
        rental_revenue = int(w['rrev'])
        rental_cost    = int(w['rcost'])                 # dated rental repair expenses (RepairLog)
        rental_profit  = rental_revenue - rental_cost
        sale_revenue   = int(w['srev'])
        sale_cost      = int(w['scogs'])
        sale_profit    = int(w['sprof'])
        accessory_revenue = int(w['arev'])
        accessory_damage_loss = int(w['aloss'])
        accessory_cost    = int(w['acogs']) + accessory_damage_loss   # COGS + write-off loss
        accessory_profit  = int(w['aprof'])
        store_cost_total = int(w['scost'])

        total_revenue = rental_revenue + sale_revenue + accessory_revenue
        total_profit  = rental_profit + sale_profit + accessory_profit
        total_cost    = rental_cost + sale_cost + accessory_cost
        net_profit    = total_profit - store_cost_total

        # ── inventory snapshot (current state, not date-scoped) ──
        in_stock     = [c for c in sale_cameras if c.sale_state == 'stock']
        fixing_units = [c for c in sale_cameras if c.sale_state == 'fixing']
        unfix_units  = [c for c in sale_cameras if c.sale_state == 'unfixable']
        inventory_value = int(sum(c.inventory_value for c in sale_cameras))
        sold_in_period = w['sold']
        top_profit = sorted(sold_in_period, key=lambda c: c.profit, reverse=True)[:8]
        aging = [{'cam': c, 'days': (now - c.date_in).days if c.date_in else None} for c in in_stock]
        aging.sort(key=lambda x: (-1 if x['days'] is None else x['days']), reverse=True)
        aging = aging[:8]

        summary = {
            'rental_revenue': rental_revenue, 'rental_cost': rental_cost, 'rental_profit': rental_profit,
            'sale_revenue': sale_revenue, 'sale_cost': sale_cost, 'sale_profit': sale_profit,
            'accessory_revenue': accessory_revenue, 'accessory_cost': accessory_cost, 'accessory_profit': accessory_profit,
            'accessory_damage_loss': accessory_damage_loss,
            'total_revenue': total_revenue, 'total_cost': total_cost, 'total_profit': total_profit,
            'inventory_value': inventory_value,
            'store_cost_total': store_cost_total, 'net_profit': net_profit,
            'units_sold': len(sold_in_period), 'in_stock_count': len(in_stock),
            'broken_count': len(fixing_units) + len(unfix_units), 'unfixable_count': len(unfix_units),
            'sold_count': len(sold_in_period), 'camera_count': len(cameras),
        }
        period_labels = {'day': '14 ngày', 'month': '12 tháng', 'year': '5 năm', 'all': 'năm'}
        if period == 'all':
            anchor_label, anchor_value = 'Toàn thời gian', ''
            prev_anchor = next_anchor = ''
            is_current = True                         # hides "next" + "về hiện tại"
        else:
            anchor_labels = {
                'day':   'Ngày ' + astart.strftime('%d/%m/%Y'),
                'month': 'Tháng ' + astart.strftime('%m/%Y'),
                'year':  'Năm ' + astart.strftime('%Y'),
            }
            anchor_values = {'day': astart.strftime('%Y-%m-%d'),
                             'month': astart.strftime('%Y-%m'), 'year': astart.strftime('%Y')}
            prev_dt = self._shift(period, anchor, 1)
            next_dt = self._shift(period, anchor, -1)
            prev_values = {'day': prev_dt.strftime('%Y-%m-%d'),
                           'month': prev_dt.strftime('%Y-%m'), 'year': prev_dt.strftime('%Y')}
            next_values = {'day': next_dt.strftime('%Y-%m-%d'),
                           'month': next_dt.strftime('%Y-%m'), 'year': next_dt.strftime('%Y')}
            anchor_label, anchor_value = anchor_labels[period], anchor_values[period]
            prev_anchor, next_anchor = prev_values[period], next_values[period]
            is_current = (astart <= now < aend)
        return self.render('admin/finance.html', summary=summary, cameras=cameras,
                           rent_cameras=rent_cameras, sale_cameras=sale_cameras,
                           period=period, period_label=period_labels[period],
                           anchor_label=anchor_label, anchor_value=anchor_value,
                           prev_anchor=prev_anchor, next_anchor=next_anchor,
                           is_current=is_current,
                           rental_series=rental_series, sale_rev_series=sale_rev_series,
                           sale_profit_series=sale_profit_series, store_cost_series=store_cost_series,
                           top_profit=top_profit, aging=aging)

    def _resolve_window(self):
        """Parse ?period + ?anchor into (period, start, end, label) — mirrors index()."""
        period = request.args.get('period', 'month')
        if period not in ('day', 'month', 'year', 'all'):
            period = 'month'
        now = datetime.now()
        anchor_raw = (request.args.get('anchor') or '').strip()
        anchor = parse_dt(anchor_raw)
        if not anchor:
            if period == 'year' and anchor_raw.isdigit():
                anchor = datetime(int(anchor_raw), 1, 1)
            elif period == 'month' and len(anchor_raw) >= 7 and anchor_raw[:4].isdigit():
                try:
                    anchor = datetime(int(anchor_raw[:4]), int(anchor_raw[5:7]), 1)
                except ValueError:
                    anchor = None
        if not anchor:
            anchor = now
        if period == 'all':
            return period, datetime(1970, 1, 1), datetime(2999, 1, 1), 'Toàn thời gian', now
        astart, aend = self._bucket(period, anchor)
        label = {'day': 'Ngày ' + astart.strftime('%d/%m/%Y'),
                 'month': 'Tháng ' + astart.strftime('%m/%Y'),
                 'year': 'Năm ' + astart.strftime('%Y')}[period]
        return period, astart, aend, label, now

    @expose('/export')
    def export(self):
        """Download the finance dashboard + current inventory as a formatted .xlsx."""
        if not HAS_OPENPYXL:
            return jsonify({'ok': False, 'error': 'Thiếu thư viện openpyxl trên máy chủ.'}), 500
        period, astart, aend, label, now = self._resolve_window()
        cameras = Camera.query.order_by(Camera.brand, Camera.name).all()
        sale_cameras = [c for c in cameras if c.type == 'Sale']
        bookings = RentalBooking.query.all()
        store_costs = StoreCost.query.all()
        acc_sales = AccessorySale.query.all()
        acc_damages = AccessoryDamage.query.all()

        repair_logs = RepairLog.query.all()
        in_win = lambda d: d and astart <= d < aend
        sold = [c for c in sale_cameras if c.sale_state in ('sold', 'deposit') and in_win(c.sold_date)]
        rrev = sum(int(b.total_price or 0) for b in bookings if in_win(b.start_date))
        rcost = sum(int(l.amount or 0) for l in repair_logs if in_win(l.created_at))
        srev = sum(int(c.sale_revenue) for c in sold)
        scogs = sum(int(c.cost) for c in sold)
        scost = sum(int(x.amount or 0) for x in store_costs if in_win(x.cost_date))
        asales = [s for s in acc_sales if in_win(s.sale_date)]
        arev = sum(s.revenue for s in asales)
        # accessory cost = COGS of sold units + write-off loss of damaged units (phụ kiện hư)
        aloss = sum(x.loss for x in acc_damages if in_win(x.damage_date))
        acogs = sum(s.cost_total for s in asales) + aloss
        inv = [c for c in sale_cameras if c.sale_state in ('stock', 'processing', 'fixing')]
        inv_value = int(sum(c.inventory_value for c in sale_cameras))
        total_rev = rrev + srev + arev
        net = (rrev - rcost) + (srev - scogs) + (arev - acogs) - scost

        wb = Workbook()
        head_fill = PatternFill('solid', fgColor='B91C1C')
        head_font = Font(bold=True, color='FFFFFF')
        money_fmt = '#,##0'

        def style_header(ws, ncols):
            for ci in range(1, ncols + 1):
                cell = ws.cell(row=1, column=ci)
                cell.fill = head_fill
                cell.font = head_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        def autosize(ws, widths):
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

        # ── Sheet 1: Tổng quan ──
        ws = wb.active
        ws.title = 'Tổng quan'
        ws.append(['tintus.digicam — Báo cáo tài chính'])
        ws.append(['Kỳ báo cáo', label])
        ws.append(['Xuất lúc', now.strftime('%d/%m/%Y %H:%M')])
        ws.append([])
        ws.append(['Khoản mục', 'Doanh thu', 'Chi phí (giá vốn)', 'Lợi nhuận'])
        hdr_row = ws.max_row
        ws.append(['Cho thuê', rrev, rcost, rrev - rcost])
        ws.append(['Bán máy', srev, scogs, srev - scogs])
        ws.append(['Phụ kiện', arev, acogs, arev - acogs])
        ws.append(['Tổng cộng', total_rev, rcost + scogs + acogs, (rrev - rcost) + (srev - scogs) + (arev - acogs)])
        ws.append([])
        ws.append(['Chi phí tiệm (kỳ này)', scost])
        ws.append(['Lợi nhuận ròng', net])
        ws.append(['Giá trị tồn kho (hiện tại)', inv_value])
        ws.append(['Số máy bán trong kỳ', len(sold)])
        ws.append(['Số máy còn trong kho', len(inv)])
        for r in range(hdr_row, hdr_row + 4):
            for c in (2, 3, 4):
                ws.cell(row=r, column=c).number_format = money_fmt
        for r in range(hdr_row + 6, hdr_row + 9):
            ws.cell(row=r, column=2).number_format = money_fmt
        for ci in range(1, 5):
            ws.cell(row=hdr_row, column=ci).fill = head_fill
            ws.cell(row=hdr_row, column=ci).font = head_font
        ws['A1'].font = Font(bold=True, size=13, color='B91C1C')
        autosize(ws, [30, 18, 18, 18])

        # ── Sheet 2: Đã bán trong kỳ ──
        ws2 = wb.create_sheet('Đã bán trong kỳ')
        ws2.append(['Tên máy', 'Hãng', 'Ngày bán', 'Trạng thái', 'Giá bán', 'Đã cọc',
                    'Giá nhập', 'Phí sửa', 'Lợi nhuận', 'Khách', 'SĐT', 'Nguồn'])
        style_header(ws2, 12)
        for c in sorted(sold, key=lambda x: x.sold_date or now, reverse=True):
            ws2.append([c.name, c.brand or '', c.sold_date.strftime('%d/%m/%Y') if c.sold_date else '',
                        c.state_label, int(c.sold_price or 0), int(c.deposit_amount or 0),
                        int(c.import_cost or 0), int(c.repair_cost or 0), int(c.profit),
                        c.sold_to or '', c.sold_phone or '', c.sold_source or ''])
        for r in range(2, ws2.max_row + 1):
            for ci in (5, 6, 7, 8, 9):
                ws2.cell(row=r, column=ci).number_format = money_fmt
        autosize(ws2, [26, 12, 13, 14, 14, 12, 12, 11, 14, 16, 13, 12])

        # ── Sheet 3: Tồn kho hiện tại ──
        ws3 = wb.create_sheet('Tồn kho hiện tại')
        ws3.append(['Tên máy', 'Hãng', 'Dòng máy', 'Trạng thái', 'Giá nhập', 'Ngày về', 'Số ngày tồn'])
        style_header(ws3, 7)
        for c in inv:
            days = (now - c.date_in).days if c.date_in else ''
            ws3.append([c.name, c.brand or '', c.category or '', c.state_label,
                        int(c.import_cost or 0), c.date_in.strftime('%d/%m/%Y') if c.date_in else '', days])
        for r in range(2, ws3.max_row + 1):
            ws3.cell(row=r, column=5).number_format = money_fmt
        autosize(ws3, [26, 12, 14, 14, 12, 13, 12])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        stamp = now.strftime('%Y%m%d')
        return send_file(buf, as_attachment=True, download_name=f'baocao-taichinh-{stamp}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


class StoreCostView(AdminOnlyView):
    """Shop overhead / operating costs (Chi phí tiệm) — rent, ads, shipping…"""

    CATEGORIES = ['Tiền nhà', 'Quảng cáo', 'Vận chuyển', 'Phụ kiện', 'Lương', 'Điện nước', 'Sửa chữa', 'Khác']

    @expose('/')
    def index(self):
        costs = StoreCost.query.order_by(StoreCost.cost_date.desc(), StoreCost.id.desc()).all()
        by_cat = {}
        for c in costs:
            by_cat[c.category or 'Khác'] = by_cat.get(c.category or 'Khác', 0) + (c.amount or 0)
        by_cat = sorted(by_cat.items(), key=lambda kv: kv[1], reverse=True)
        total = int(sum(c.amount or 0 for c in costs))
        now = datetime.now()
        month_total = int(sum((c.amount or 0) for c in costs
                              if c.cost_date and c.cost_date.year == now.year and c.cost_date.month == now.month))
        return self.render('admin/chiphi.html',
                           costs=[serialize_cost(c) for c in costs],
                           by_cat=by_cat, total=total, month_total=month_total,
                           categories=self.CATEGORIES, today=now.strftime('%Y-%m-%d'))

    @expose('/add', methods=['POST'])
    def add(self):
        d = request.get_json(silent=True) or {}
        try:
            amount = max(0, int(d.get('amount') or 0))
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'Số tiền không hợp lệ'}), 400
        c = StoreCost(category=(d.get('category') or 'Khác').strip(),
                      note=(d.get('note') or '').strip(), amount=amount,
                      cost_date=parse_dt(d.get('cost_date')) or datetime.now())
        db.session.add(c)
        db.session.commit()
        return jsonify({'ok': True, 'cost': serialize_cost(c)})

    @expose('/update', methods=['POST'])
    def update(self):
        d = request.get_json(silent=True) or {}
        c = db.session.get(StoreCost, d.get('id'))
        if not c:
            return jsonify({'ok': False, 'error': 'Không tìm thấy'}), 404
        f, v = d.get('field'), d.get('value')
        if f == 'amount':
            try:
                c.amount = max(0, int(v or 0))
            except (TypeError, ValueError):
                return jsonify({'ok': False, 'error': 'Số tiền không hợp lệ'}), 400
        elif f == 'category':
            c.category = (str(v) if v else 'Khác').strip()
        elif f == 'note':
            c.note = (str(v) if v else '').strip()
        elif f == 'cost_date':
            c.cost_date = parse_dt(v) or c.cost_date
        else:
            return jsonify({'ok': False, 'error': 'Trường không hợp lệ'}), 400
        db.session.commit()
        return jsonify({'ok': True})

    @expose('/delete', methods=['POST'])
    def delete(self):
        d = request.get_json(silent=True) or {}
        c = db.session.get(StoreCost, d.get('id'))
        if not c:
            return jsonify({'ok': False, 'error': 'Không tìm thấy'}), 404
        db.session.delete(c)
        db.session.commit()
        return jsonify({'ok': True})


class AccessoriesView(StaffView):
    """Phụ kiện / hàng hoá khác — sold by quantity (lens, pin, thẻ nhớ…).
    Employees can check stock and sell (POS); catalog CRUD and costs are admin-only."""

    EDITABLE = {'name': str, 'category': str, 'note': str,
                'cost': int, 'price': int, 'stock': int}

    @expose('/')
    def index(self):
        items = Accessory.query.order_by(Accessory.category, Accessory.name).all()
        total_stock  = int(sum(a.stock or 0 for a in items))
        stock_value  = int(sum(a.stock_value for a in items))
        now = datetime.now()
        month_sales = [s for s in AccessorySale.query.all()
                       if s.sale_date and s.sale_date.year == now.year and s.sale_date.month == now.month]
        month_rev = int(sum(s.revenue for s in month_sales))
        sales = (AccessorySale.query.order_by(AccessorySale.sale_date.desc()).limit(300).all())
        damages = (AccessoryDamage.query.order_by(AccessoryDamage.damage_date.desc()).limit(300).all())
        month_damages = [x for x in damages
                         if x.damage_date and x.damage_date.year == now.year and x.damage_date.month == now.month]
        month_loss = int(sum(x.loss for x in month_damages))
        return self.render('admin/phukien.html',
                           accessories=[redact_accessory(serialize_accessory(a)) for a in items],
                           sales=[redact_accessory_sale(serialize_accessory_sale(s)) for s in sales],
                           damages=[redact_accessory_damage(serialize_accessory_damage(x)) for x in damages],
                           categories=ACCESSORY_CATEGORIES,
                           total_stock=total_stock,
                           stock_value=stock_value if is_admin() else 0,
                           item_count=len(items), month_rev=month_rev,
                           month_loss=month_loss if is_admin() else 0,
                           today=now.strftime('%Y-%m-%d'),
                           this_month=now.strftime('%Y-%m'), this_year=now.strftime('%Y'))

    @expose('/add', methods=['POST'])
    def add(self):
        if not is_admin():
            return _forbidden('Chỉ quản lý được thêm phụ kiện vào kho.')
        d = request.get_json(silent=True) or {}
        name = (d.get('name') or '').strip()
        if not name:
            return jsonify({'ok': False, 'error': 'Tên phụ kiện bắt buộc'}), 400
        try:
            a = Accessory(name=name, category=(d.get('category') or '').strip(),
                          cost=max(0, int(d.get('cost') or 0)),
                          price=max(0, int(d.get('price') or 0)),
                          stock=max(0, int(d.get('stock') or 0)),
                          note=(d.get('note') or '').strip())
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'Giá trị không hợp lệ'}), 400
        db.session.add(a)
        db.session.commit()
        return jsonify({'ok': True, 'accessory': serialize_accessory(a)})

    @expose('/update', methods=['POST'])
    def update(self):
        if not is_admin():
            return _forbidden('Chỉ quản lý được sửa thông tin phụ kiện.')
        d = request.get_json(silent=True) or {}
        a = db.session.get(Accessory, d.get('id'))
        if not a:
            return jsonify({'ok': False, 'error': 'Không tìm thấy'}), 404
        field, value = d.get('field'), d.get('value')
        if field not in self.EDITABLE:
            return jsonify({'ok': False, 'error': 'Trường không hợp lệ'}), 400
        caster = self.EDITABLE[field]
        try:
            cast = max(0, int(value or 0)) if caster is int else (str(value) if value is not None else '').strip()
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'Giá trị không hợp lệ'}), 400
        setattr(a, field, cast)
        db.session.commit()
        return jsonify({'ok': True, 'accessory': serialize_accessory(a)})

    @expose('/sell', methods=['POST'])
    def sell(self):
        """Sell N units at a price → decrement stock, log the sale."""
        d = request.get_json(silent=True) or {}
        a = db.session.get(Accessory, d.get('id'))
        if not a:
            return jsonify({'ok': False, 'error': 'Không tìm thấy'}), 404
        try:
            qty   = int(d.get('quantity') or 0)
            price = max(0, int(d.get('price') or 0))
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'Số lượng / giá không hợp lệ'}), 400
        if qty < 1:
            return jsonify({'ok': False, 'error': 'Số lượng phải ≥ 1'}), 400
        if price <= 0:
            return jsonify({'ok': False, 'error': 'Giá bán phải lớn hơn 0đ.'}), 400
        if qty > (a.stock or 0):
            return jsonify({'ok': False, 'error': f'Chỉ còn {a.stock or 0} cái trong kho'}), 400
        # Employee discount floor: listed price minus the admin-set max %.
        if not is_admin() and (a.price or 0) > 0:
            pct = setting_float('max_discount_pct')
            floor_price = int((a.price or 0) * (100 - pct) / 100)
            if price < floor_price:
                return _forbidden(f'Giá bán thấp hơn mức cho phép (giảm tối đa {pct:g}%). Liên hệ quản lý.')
        a.stock -= qty
        rec = AccessorySale(accessory_id=a.id, name=a.name, quantity=qty,
                            unit_price=price, unit_cost=int(a.cost or 0),
                            note=(d.get('note') or '').strip(),
                            customer_name=(d.get('customer_name') or '').strip(),
                            phone=(d.get('phone') or '').strip(),
                            staff=current_staff_user(),
                            sale_date=datetime.now())
        db.session.add(rec)
        db.session.commit()
        return jsonify({'ok': True, 'accessory': redact_accessory(serialize_accessory(a)),
                        'sold': qty, 'revenue': rec.revenue,
                        'sale': redact_accessory_sale(serialize_accessory_sale(rec))})

    @expose('/sell-service', methods=['POST'])
    def sell_service(self):
        """Record a free-form service / other-item sale (dịch vụ khác) that isn't tied
        to a stocked accessory — e.g. vệ sinh máy, phí ship, dán màn hình. Stored as an
        AccessorySale with no accessory_id so it feeds Finance like any other sale.
        Cost (COGS) is optional and admin-only; employees just enter a price."""
        d = request.get_json(silent=True) or {}
        name = (d.get('name') or '').strip()
        if not name:
            return jsonify({'ok': False, 'error': 'Nhập tên dịch vụ / mặt hàng'}), 400
        try:
            qty   = max(1, int(d.get('quantity') or 1))
            price = max(0, int(d.get('price') or 0))
            cost  = max(0, int(d.get('cost') or 0)) if is_admin() else 0
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'Số lượng / giá không hợp lệ'}), 400
        if price <= 0:
            return jsonify({'ok': False, 'error': 'Giá bán phải lớn hơn 0đ.'}), 400
        rec = AccessorySale(accessory_id=None, name=name, quantity=qty,
                            unit_price=price, unit_cost=cost,
                            note=(d.get('note') or '').strip(),
                            customer_name=(d.get('customer_name') or '').strip(),
                            phone=(d.get('phone') or '').strip(),
                            staff=current_staff_user(),
                            sale_date=datetime.now())
        db.session.add(rec)
        db.session.commit()
        return jsonify({'ok': True, 'sold': qty, 'revenue': rec.revenue,
                        'sale': redact_accessory_sale(serialize_accessory_sale(rec))})

    # ── Sales-history editing (Lịch sử bán) ──────────────────────────────────
    SALE_EDITABLE = {'quantity': int, 'unit_price': int, 'note': str,
                     'customer_name': str, 'phone': str}

    @expose('/sale/update', methods=['POST'])
    def sale_update(self):
        """Inline-edit a logged accessory sale. Editing quantity re-syncs stock."""
        d = request.get_json(silent=True) or {}
        rec = db.session.get(AccessorySale, d.get('id'))
        if not rec:
            return jsonify({'ok': False, 'error': 'Không tìm thấy giao dịch'}), 404
        field, value = d.get('field'), d.get('value')
        if field not in self.SALE_EDITABLE:
            return jsonify({'ok': False, 'error': 'Trường không hợp lệ'}), 400
        # Employees may fix customer info / notes; money & quantity edits are admin-only.
        if not is_admin() and field not in ('note', 'customer_name', 'phone'):
            return _forbidden('Chỉ quản lý được sửa số lượng / giá của giao dịch đã ghi.')
        if field == 'quantity':
            try:
                new_qty = max(1, int(value or 1))
            except (TypeError, ValueError):
                return jsonify({'ok': False, 'error': 'Số lượng không hợp lệ'}), 400
            delta = new_qty - (rec.quantity or 0)          # extra units to pull from stock
            acc = db.session.get(Accessory, rec.accessory_id) if rec.accessory_id else None
            if acc and delta > 0 and delta > (acc.stock or 0):
                return jsonify({'ok': False, 'error': f'Chỉ còn {acc.stock or 0} cái trong kho'}), 400
            if acc:
                acc.stock = max(0, (acc.stock or 0) - delta)
            rec.quantity = new_qty
        elif field == 'unit_price':
            try:
                rec.unit_price = max(0, int(value or 0))
            except (TypeError, ValueError):
                return jsonify({'ok': False, 'error': 'Giá không hợp lệ'}), 400
        else:
            setattr(rec, field, (str(value) if value is not None else '').strip())
        db.session.commit()
        acc = db.session.get(Accessory, rec.accessory_id) if rec.accessory_id else None
        return jsonify({'ok': True, 'sale': redact_accessory_sale(serialize_accessory_sale(rec)),
                        'accessory': redact_accessory(serialize_accessory(acc)) if acc else None})

    @expose('/sale/delete', methods=['POST'])
    def sale_delete(self):
        """Delete a logged accessory sale and return its units to stock."""
        if not is_admin():
            return _forbidden('Chỉ quản lý được xoá giao dịch đã ghi.')
        d = request.get_json(silent=True) or {}
        rec = db.session.get(AccessorySale, d.get('id'))
        if not rec:
            return jsonify({'ok': False, 'error': 'Không tìm thấy giao dịch'}), 404
        acc = db.session.get(Accessory, rec.accessory_id) if rec.accessory_id else None
        if acc:
            acc.stock = (acc.stock or 0) + (rec.quantity or 0)
        db.session.delete(rec)
        db.session.commit()
        return jsonify({'ok': True, 'accessory': serialize_accessory(acc) if acc else None})

    @expose('/delete', methods=['POST'])
    def delete(self):
        if not is_admin():
            return _forbidden('Chỉ quản lý được xoá phụ kiện khỏi kho.')
        d = request.get_json(silent=True) or {}
        a = db.session.get(Accessory, d.get('id'))
        if not a:
            return jsonify({'ok': False, 'error': 'Không tìm thấy'}), 404
        db.session.delete(a)
        db.session.commit()
        return jsonify({'ok': True})

    # ── Phụ kiện hư (damaged / write-off) ─────────────────────────────────────
    @expose('/damage', methods=['POST'])
    def damage(self):
        """Report N units of an accessory as broken/unsellable → remove them from
        sellable stock and log the loss (staff + admin)."""
        d = request.get_json(silent=True) or {}
        a = db.session.get(Accessory, d.get('id'))
        if not a:
            return jsonify({'ok': False, 'error': 'Không tìm thấy'}), 404
        try:
            qty = int(d.get('quantity') or 0)
        except (TypeError, ValueError):
            return jsonify({'ok': False, 'error': 'Số lượng không hợp lệ'}), 400
        if qty < 1:
            return jsonify({'ok': False, 'error': 'Số lượng phải ≥ 1'}), 400
        if qty > (a.stock or 0):
            return jsonify({'ok': False, 'error': f'Chỉ còn {a.stock or 0} cái trong kho'}), 400
        a.stock -= qty
        rec = AccessoryDamage(accessory_id=a.id, name=a.name, quantity=qty,
                              unit_cost=int(a.cost or 0),
                              note=(d.get('note') or '').strip(),
                              staff=current_staff_user(),
                              damage_date=datetime.now())
        db.session.add(rec)
        db.session.commit()
        return jsonify({'ok': True, 'accessory': redact_accessory(serialize_accessory(a)),
                        'damaged': qty,
                        'damage': redact_accessory_damage(serialize_accessory_damage(rec))})

    @expose('/damage/delete', methods=['POST'])
    def damage_delete(self):
        """Undo a damage write-off and return its units to stock (admin only)."""
        if not is_admin():
            return _forbidden('Chỉ quản lý được xoá ghi nhận phụ kiện hư.')
        d = request.get_json(silent=True) or {}
        rec = db.session.get(AccessoryDamage, d.get('id'))
        if not rec:
            return jsonify({'ok': False, 'error': 'Không tìm thấy'}), 404
        acc = db.session.get(Accessory, rec.accessory_id) if rec.accessory_id else None
        if acc:
            acc.stock = (acc.stock or 0) + (rec.quantity or 0)
        db.session.delete(rec)
        db.session.commit()
        return jsonify({'ok': True, 'accessory': serialize_accessory(acc) if acc else None})


def serialize_lead(l):
    return {
        'id': l.id, 'customer_name': l.customer_name or '', 'phone': l.phone or '',
        'notes': l.notes or '', 'source': l.source or '', 'status': l.status or 'new',
        'want_start': l.start_date.strftime('%d/%m/%Y') if l.start_date else '',
        'want_end':   l.end_date.strftime('%d/%m/%Y') if l.end_date else '',
        'created': l.created_at.strftime('%d/%m/%Y %H:%M') if l.created_at else '',
        'created_iso': l.created_at.strftime('%Y-%m-%d') if l.created_at else '',
    }


class LeadView(StaffView):
    """Khách quan tâm — inquiries captured from the public homepage form."""
    STATUSES = ('new', 'contacted', 'done')
    EDITABLE = ('customer_name', 'phone', 'notes', 'status')

    @expose('/')
    def index(self):
        leads = Lead.query.order_by(Lead.created_at.desc(), Lead.id.desc()).all()
        counts = {s: 0 for s in self.STATUSES}
        for l in leads:
            counts[l.status if l.status in counts else 'new'] += 1
        return self.render('admin/leads.html',
                           leads=[serialize_lead(l) for l in leads],
                           counts=counts, total=len(leads))

    @expose('/update', methods=['POST'])
    def update(self):
        d = request.get_json(silent=True) or {}
        lead = db.session.get(Lead, d.get('id'))
        if not lead:
            return jsonify({'ok': False, 'error': 'Không tìm thấy'}), 404
        field, value = d.get('field'), d.get('value')
        if field not in self.EDITABLE:
            return jsonify({'ok': False, 'error': 'Trường không hợp lệ'}), 400
        if field == 'status' and value not in self.STATUSES:
            return jsonify({'ok': False, 'error': 'Trạng thái không hợp lệ'}), 400
        setattr(lead, field, (str(value) if value is not None else '').strip())
        db.session.commit()
        return jsonify({'ok': True})

    @expose('/delete', methods=['POST'])
    def delete(self):
        d = request.get_json(silent=True) or {}
        lead = db.session.get(Lead, d.get('id'))
        if not lead:
            return jsonify({'ok': False, 'error': 'Không tìm thấy'}), 404
        db.session.delete(lead)
        db.session.commit()
        return jsonify({'ok': True})


class SheetView(AdminOnlyView):
    """A free-form Google-Sheets-style scratch sheet (admin only — may hold
    financial notes)."""

    def _get_sheet(self):
        sheet = Sheet.query.first()
        if not sheet:
            blank = [['' for _ in range(8)] for _ in range(30)]
            sheet = Sheet(name='Sổ tay', data_json=json.dumps(blank))
            db.session.add(sheet)
            db.session.commit()
        return sheet

    @expose('/')
    def index(self):
        return self.render('admin/sheet.html')

    @expose('/load')
    def load(self):
        sheet = self._get_sheet()
        return jsonify({'name': sheet.name, 'data': sheet.data,
                        'updated_at': sheet.updated_at.isoformat() if sheet.updated_at else None})

    @expose('/save', methods=['POST'])
    def save(self):
        d = request.get_json(silent=True) or {}
        data = d.get('data')
        if not isinstance(data, list):
            return jsonify({'ok': False, 'error': 'Dữ liệu không hợp lệ'}), 400
        # Coerce to a clean 2D array of strings.
        clean = [[('' if c is None else str(c)) for c in (row if isinstance(row, list) else [])]
                 for row in data]
        sheet = self._get_sheet()
        if d.get('name'):
            sheet.name = str(d['name']).strip()[:100]
        sheet.data_json = json.dumps(clean, ensure_ascii=False)
        db.session.commit()
        return jsonify({'ok': True, 'updated_at': sheet.updated_at.isoformat() if sheet.updated_at else None})


class ActivityView(AdminOnlyView):
    """Nhật ký — chronological log of every admin action, with undo / redo.
    Admin only: change details include import costs and profits."""

    @expose('/')
    def index(self):
        logs = ActivityLog.query.order_by(ActivityLog.id.desc()).limit(300).all()
        rows = []
        for l in logs:
            try:
                changes = _act_loads(l.changes_json)
            except Exception:
                changes = []
            rows.append({
                'id': l.id, 'label': l.label or '(hoạt động)', 'undone': bool(l.undone),
                'when': l.created_at.strftime('%d/%m/%Y %H:%M:%S') if l.created_at else '',
                'count': len(changes), 'details': _act_detail_lines(changes),
            })
        return self.render('admin/activity.html', rows=rows,
                           can_undo=any(not r['undone'] for r in rows),
                           can_redo=any(r['undone'] for r in rows))

    @expose('/undo', methods=['POST'])
    def undo(self):
        log = (ActivityLog.query.filter_by(undone=False)
               .order_by(ActivityLog.id.desc()).first())
        if not log:
            return jsonify({'ok': False, 'error': 'Không có gì để hoàn tác'}), 400
        try:
            g._act_suspend = True
            _act_reverse(_act_loads(log.changes_json))
            log.undone = True
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            return jsonify({'ok': False, 'error': f'Không hoàn tác được: {exc}'}), 500
        finally:
            g._act_suspend = False
        return jsonify({'ok': True, 'label': log.label})

    @expose('/redo', methods=['POST'])
    def redo(self):
        log = (ActivityLog.query.filter_by(undone=True)
               .order_by(ActivityLog.id.asc()).first())
        if not log:
            return jsonify({'ok': False, 'error': 'Không có gì để làm lại'}), 400
        try:
            g._act_suspend = True
            _act_apply(_act_loads(log.changes_json))
            log.undone = False
            db.session.commit()
        except Exception as exc:
            db.session.rollback()
            return jsonify({'ok': False, 'error': f'Không làm lại được: {exc}'}), 500
        finally:
            g._act_suspend = False
        return jsonify({'ok': True, 'label': log.label})

    @expose('/clear', methods=['POST'])
    def clear(self):
        """Wipe the whole history (does not touch the data itself)."""
        try:
            g._act_suspend = True
            ActivityLog.query.delete()
            db.session.commit()
        finally:
            g._act_suspend = False
        return jsonify({'ok': True})


def serialize_employee(e):
    return {'id': e.id, 'username': e.username, 'display_name': e.display_name or '',
            'role': e.role or 'employee', 'active': bool(e.active),
            'created': e.created_at.strftime('%d/%m/%Y') if e.created_at else ''}


class EmployeesView(AdminOnlyView):
    """Nhân viên — staff accounts (create / edit / disable / reset password)
    plus per-employee performance (sales volume, rental contracts processed)."""

    @expose('/')
    def index(self):
        emps = Employee.query.order_by(Employee.created_at.asc()).all()
        now = datetime.now()
        month_start = datetime(now.year, now.month, 1)

        def perf(username, since=None):
            bq = RentalBooking.query.filter(RentalBooking.staff == username)
            if since:
                bq = bq.filter(RentalBooking.start_date >= since)
            bookings = bq.all()
            cams = Camera.query.filter(Camera.sold_by == username,
                                       Camera.sale_state.in_(Camera.SOLD_STATES)).all()
            if since:
                cams = [c for c in cams if c.sold_date and c.sold_date >= since]
            accq = AccessorySale.query.filter(AccessorySale.staff == username)
            accs = [s for s in accq.all() if not since or (s.sale_date and s.sale_date >= since)]
            return {
                'rentals': len(bookings),
                'rental_total': int(sum(b.total_price or 0 for b in bookings)),
                'cam_sales': len(cams),
                'cam_total': int(sum(c.sold_price or 0 for c in cams)),
                'acc_sales': len(accs),
                'acc_total': int(sum(s.revenue for s in accs)),
            }

        rows = []
        for e in emps:
            rows.append(serialize_employee(e) | {
                'month': perf(e.username, month_start),
                'all':   perf(e.username),
            })
        return self.render('admin/employees.html', employees=rows,
                           this_month=now.strftime('%m/%Y'))

    @expose('/add', methods=['POST'])
    def add(self):
        d = request.get_json(silent=True) or {}
        username = (d.get('username') or '').strip().lower()
        password = d.get('password') or ''
        if not username or len(username) < 3:
            return jsonify({'ok': False, 'error': 'Tên đăng nhập tối thiểu 3 ký tự'}), 400
        if len(password) < 6:
            return jsonify({'ok': False, 'error': 'Mật khẩu tối thiểu 6 ký tự'}), 400
        if username == ADMIN_USERNAME.lower() or \
           Employee.query.filter(db.func.lower(Employee.username) == username).first():
            return jsonify({'ok': False, 'error': 'Tên đăng nhập đã tồn tại'}), 400
        role = d.get('role') if d.get('role') in ('admin', 'employee') else 'employee'
        e = Employee(username=username, password_hash=generate_password_hash(password),
                     display_name=(d.get('display_name') or '').strip() or username,
                     role=role, active=True)
        db.session.add(e)
        db.session.commit()
        return jsonify({'ok': True, 'employee': serialize_employee(e)})

    @expose('/update', methods=['POST'])
    def update(self):
        d = request.get_json(silent=True) or {}
        e = db.session.get(Employee, d.get('id'))
        if not e:
            return jsonify({'ok': False, 'error': 'Không tìm thấy nhân viên'}), 404
        field, value = d.get('field'), d.get('value')
        if field == 'display_name':
            e.display_name = (str(value) if value else '').strip()[:100]
        elif field == 'role':
            if value not in ('admin', 'employee'):
                return jsonify({'ok': False, 'error': 'Vai trò không hợp lệ'}), 400
            e.role = value
        elif field == 'active':
            e.active = bool(value) if isinstance(value, bool) else str(value).lower() in ('1', 'true', 'on')
        else:
            return jsonify({'ok': False, 'error': 'Trường không hợp lệ'}), 400
        db.session.commit()
        return jsonify({'ok': True, 'employee': serialize_employee(e)})

    @expose('/reset-password', methods=['POST'])
    def reset_password(self):
        d = request.get_json(silent=True) or {}
        e = db.session.get(Employee, d.get('id'))
        if not e:
            return jsonify({'ok': False, 'error': 'Không tìm thấy nhân viên'}), 404
        password = d.get('password') or ''
        if len(password) < 6:
            return jsonify({'ok': False, 'error': 'Mật khẩu tối thiểu 6 ký tự'}), 400
        e.password_hash = generate_password_hash(password)
        db.session.commit()
        return jsonify({'ok': True})

    @expose('/delete', methods=['POST'])
    def delete(self):
        d = request.get_json(silent=True) or {}
        e = db.session.get(Employee, d.get('id'))
        if not e:
            return jsonify({'ok': False, 'error': 'Không tìm thấy nhân viên'}), 404
        db.session.delete(e)
        db.session.commit()
        return jsonify({'ok': True})


def serialize_customer(c):
    return {'id': c.id, 'name': c.name or '', 'phone': c.phone or '',
            'email': c.email or '', 'address': c.address or '', 'note': c.note or '',
            'is_banned': bool(c.is_banned), 'ban_reason': c.ban_reason or '',
            'created_by': c.created_by or '',
            'created': c.created_at.strftime('%d/%m/%Y') if c.created_at else ''}


class CustomersView(StaffView):
    """Khách hàng — customer profiles with full purchase & rental history
    (linked by phone). Banning / blacklisting is admin-only."""

    EDITABLE = ('name', 'phone', 'email', 'address', 'note')

    @expose('/')
    def index(self):
        customers = Customer.query.order_by(Customer.created_at.desc(), Customer.id.desc()).all()
        # quick counts by phone so the list shows activity at a glance
        rental_counts, buy_counts = {}, {}
        for (phone,) in db.session.query(RentalBooking.phone).all():
            if phone:
                rental_counts[phone] = rental_counts.get(phone, 0) + 1
        for (phone,) in db.session.query(Camera.sold_phone).filter(
                Camera.sale_state.in_(Camera.SOLD_STATES)).all():
            if phone:
                buy_counts[phone] = buy_counts.get(phone, 0) + 1
        for (phone,) in db.session.query(AccessorySale.phone).all():
            if phone:
                buy_counts[phone] = buy_counts.get(phone, 0) + 1
        rows = [serialize_customer(c) | {
            'rentals': rental_counts.get(c.phone or '', 0),
            'purchases': buy_counts.get(c.phone or '', 0),
        } for c in customers]
        return self.render('admin/customers.html', customers=rows,
                           banned_count=sum(1 for c in customers if c.is_banned))

    @expose('/add', methods=['POST'])
    def add(self):
        d = request.get_json(silent=True) or {}
        name = (d.get('name') or '').strip()
        phone = (d.get('phone') or '').strip()
        if not (name or phone):
            return jsonify({'ok': False, 'error': 'Cần ít nhất tên hoặc SĐT'}), 400
        if phone and Customer.query.filter_by(phone=phone).first():
            return jsonify({'ok': False, 'error': 'Đã có hồ sơ với SĐT này'}), 400
        c = Customer(name=name, phone=phone,
                     email=(d.get('email') or '').strip(),
                     address=(d.get('address') or '').strip(),
                     note=(d.get('note') or '').strip(),
                     created_by=current_staff_user())
        db.session.add(c)
        db.session.commit()
        return jsonify({'ok': True, 'customer': serialize_customer(c)})

    @expose('/update', methods=['POST'])
    def update(self):
        d = request.get_json(silent=True) or {}
        c = db.session.get(Customer, d.get('id'))
        if not c:
            return jsonify({'ok': False, 'error': 'Không tìm thấy khách hàng'}), 404
        field, value = d.get('field'), d.get('value')
        if field in self.EDITABLE:
            setattr(c, field, (str(value) if value is not None else '').strip())
        elif field == 'is_banned':                     # blacklist — admin only
            if not is_admin():
                return _forbidden('Chỉ quản lý được chặn / bỏ chặn khách.')
            c.is_banned = bool(value) if isinstance(value, bool) else str(value).lower() in ('1', 'true', 'on')
            if not c.is_banned:
                c.ban_reason = ''
        elif field == 'ban_reason':
            if not is_admin():
                return _forbidden('Chỉ quản lý được sửa lý do chặn.')
            c.ban_reason = (str(value) if value is not None else '').strip()[:200]
        else:
            return jsonify({'ok': False, 'error': 'Trường không hợp lệ'}), 400
        db.session.commit()
        return jsonify({'ok': True, 'customer': serialize_customer(c)})

    @expose('/delete', methods=['POST'])
    def delete(self):
        if not is_admin():
            return _forbidden('Chỉ quản lý được xoá hồ sơ khách hàng.')
        d = request.get_json(silent=True) or {}
        c = db.session.get(Customer, d.get('id'))
        if not c:
            return jsonify({'ok': False, 'error': 'Không tìm thấy khách hàng'}), 404
        db.session.delete(c)
        db.session.commit()
        return jsonify({'ok': True})

    @expose('/history')
    def history(self):
        """Full rental + purchase history for one phone number (assists POS)."""
        phone = (request.args.get('phone') or '').strip()
        if not phone:
            return jsonify({'ok': False, 'error': 'Thiếu SĐT'}), 400
        rentals = (RentalBooking.query.filter_by(phone=phone)
                   .order_by(RentalBooking.start_date.desc()).limit(100).all())
        cams = (Camera.query.filter(Camera.sold_phone == phone,
                                    Camera.sale_state.in_(Camera.SOLD_STATES))
                .order_by(Camera.sold_date.desc()).limit(100).all())
        accs = (AccessorySale.query.filter_by(phone=phone)
                .order_by(AccessorySale.sale_date.desc()).limit(100).all())
        return jsonify({'ok': True,
            'rentals': [serialize_booking(b) for b in rentals],
            'cameras': [{'id': c.id, 'name': c.name, 'sold_price': int(c.sold_price or 0),
                         'state_label': c.state_label,
                         'date': c.sold_date.strftime('%d/%m/%Y') if c.sold_date else ''} for c in cams],
            'accessories': [redact_accessory_sale(serialize_accessory_sale(s)) for s in accs]})


class MyView(StaffView):
    """Cá nhân — the logged-in person's own daily/monthly numbers and
    account settings (password change for staff accounts)."""

    @expose('/')
    def index(self):
        username = current_staff_user()
        now = datetime.now()
        day_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        month_start = datetime(now.year, now.month, 1)

        def window(since):
            bookings = [b for b in RentalBooking.query.filter(RentalBooking.staff == username).all()
                        if b.start_date and b.start_date >= since]
            cams = [c for c in Camera.query.filter(Camera.sold_by == username,
                                                   Camera.sale_state.in_(Camera.SOLD_STATES)).all()
                    if c.sold_date and c.sold_date >= since]
            accs = [s for s in AccessorySale.query.filter(AccessorySale.staff == username).all()
                    if s.sale_date and s.sale_date >= since]
            return {
                'rentals': len(bookings),
                'rental_total': int(sum(b.total_price or 0 for b in bookings)),
                'cam_sales': len(cams),
                'cam_total': int(sum(c.sold_price or 0 for c in cams)),
                'acc_sales': len(accs),
                'acc_total': int(sum(s.revenue for s in accs)),
            }

        # recent transactions for the "my activity" list
        recent = []
        for b in (RentalBooking.query.filter(RentalBooking.staff == username)
                  .order_by(RentalBooking.id.desc()).limit(10).all()):
            recent.append({'kind': 'Thuê', 'label': (b.camera.name if b.camera else '—') + ' · ' + (b.customer_name or ''),
                           'amount': int(b.total_price or 0), '_ts': b.start_date,
                           'when': b.start_date.strftime('%d/%m/%Y') if b.start_date else ''})
        for c in (Camera.query.filter(Camera.sold_by == username,
                                      Camera.sale_state.in_(Camera.SOLD_STATES))
                  .order_by(Camera.sold_date.desc()).limit(10).all()):
            recent.append({'kind': 'Bán máy', 'label': c.name + (' · ' + c.sold_to if c.sold_to else ''),
                           'amount': int(c.sold_price or 0), '_ts': c.sold_date,
                           'when': c.sold_date.strftime('%d/%m/%Y') if c.sold_date else ''})
        for s in (AccessorySale.query.filter(AccessorySale.staff == username)
                  .order_by(AccessorySale.id.desc()).limit(10).all()):
            recent.append({'kind': 'Phụ kiện', 'label': f'{s.name} ×{s.quantity}' + (' · ' + s.customer_name if s.customer_name else ''),
                           'amount': s.revenue, '_ts': s.sale_date,
                           'when': s.sale_date.strftime('%d/%m/%Y') if s.sale_date else ''})
        # Sort by the real timestamp — NOT the "%d/%m/%Y" display string (which would
        # order lexicographically, e.g. 03/07 before 25/06).
        recent.sort(key=lambda r: r['_ts'] or datetime.min, reverse=True)

        is_env_admin = (username == ADMIN_USERNAME and
                        not Employee.query.filter(db.func.lower(Employee.username) == username.lower()).first())
        return self.render('admin/my.html',
                           today_stats=window(day_start), month_stats=window(month_start),
                           recent=recent[:15], is_env_admin=is_env_admin,
                           today_label=now.strftime('%d/%m/%Y'),
                           month_label=now.strftime('%m/%Y'))

    @expose('/change-password', methods=['POST'])
    def change_password(self):
        d = request.get_json(silent=True) or {}
        old, new = d.get('old') or '', d.get('new') or ''
        emp = Employee.query.filter(
            db.func.lower(Employee.username) == current_staff_user().lower()).first()
        if not emp:
            return jsonify({'ok': False, 'error': 'Tài khoản chủ tiệm đổi mật khẩu qua biến môi trường ADMIN_PASSWORD.'}), 400
        if not check_password_hash(emp.password_hash, old):
            return jsonify({'ok': False, 'error': 'Mật khẩu hiện tại không đúng'}), 400
        if len(new) < 6:
            return jsonify({'ok': False, 'error': 'Mật khẩu mới tối thiểu 6 ký tự'}), 400
        emp.password_hash = generate_password_hash(new)
        db.session.commit()
        return jsonify({'ok': True})


class SettingsView(AdminOnlyView):
    """Cài đặt — store policies, tax rate, receipt header/footer, discount
    ceiling, and a one-click database backup download."""

    @expose('/')
    def index(self):
        return self.render('admin/settings.html', settings=all_settings())

    @expose('/save', methods=['POST'])
    def save(self):
        d = request.get_json(silent=True) or {}
        for key in DEFAULT_SETTINGS:
            if key not in d:
                continue
            value = str(d.get(key) if d.get(key) is not None else '').strip()
            if key in ('tax_rate', 'max_discount_pct'):
                try:
                    value = str(max(0.0, min(100.0, float(value or 0))))
                except (TypeError, ValueError):
                    return jsonify({'ok': False, 'error': f'Giá trị "{key}" phải là số 0–100'}), 400
            row = db.session.get(Setting, key)
            if row:
                row.value = value
            else:
                db.session.add(Setting(key=key, value=value))
        db.session.commit()
        return jsonify({'ok': True, 'settings': all_settings()})

    @expose('/backup')
    def backup(self):
        """Download the SQLite database file (offline backup for accounting)."""
        path = db.engine.url.database
        if not path or not os.path.isfile(path):
            return jsonify({'ok': False, 'error': 'Không tìm thấy file cơ sở dữ liệu.'}), 404
        stamp = datetime.now().strftime('%Y%m%d-%H%M')
        return send_file(path, as_attachment=True,
                         download_name=f'camerashop-backup-{stamp}.db')


class ReceiptView(StaffView):
    """Hoá đơn — printable invoices / receipts for rentals, camera sales and
    accessory sales (browser print → paper or PDF for email)."""

    def _ctx(self):
        s = all_settings()
        return {'settings': s, 'tax_rate': setting_float('tax_rate'),
                'printed_at': datetime.now().strftime('%d/%m/%Y %H:%M'),
                'staff': current_staff()}

    @expose('/')
    def index(self):
        # no standalone receipt list — receipts are opened from a transaction
        return redirect(url_for('cameras.index'))

    @expose('/rental/<int:bid>')
    def rental(self, bid):
        b = db.session.get(RentalBooking, bid)
        if not b:
            return 'Không tìm thấy đơn thuê', 404
        days = max(1, math.ceil((b.end_date - b.start_date).total_seconds() / 86400))
        items = [{'label': f'Thuê {b.camera.name if b.camera else "máy ảnh"} ({days} ngày, '
                           f'{b.start_date.strftime("%d/%m")} → {b.end_date.strftime("%d/%m/%Y")})',
                  'qty': 1, 'amount': int(b.total_price or 0)}]
        # A security deposit is refunded on return — show it as info, don't net it off.
        note_parts = []
        if b.security_deposit:
            note_parts.append('Tiền cọc giữ máy: {:,.0f}đ (hoàn khi trả máy).'
                              .format(b.security_deposit).replace(',', '.'))
        if b.condition_out:
            note_parts.append(f'Tình trạng máy lúc giao: {b.condition_out}')
        if b.notes:
            note_parts.append(b.notes)
        return self.render('admin/receipt.html', kind='Hoá đơn thuê máy', ref=f'R{b.id:05d}',
                           customer=b.customer_name or '', phone=b.phone or '',
                           items=items, deposit=0,
                           note='\n'.join(note_parts), **self._ctx())

    @expose('/sale/<int:cid>')
    def sale(self, cid):
        c = db.session.get(Camera, cid)
        if not c or c.sale_state not in Camera.SOLD_STATES:
            return 'Không tìm thấy giao dịch bán', 404
        items = [{'label': f'{c.name}' + (f' ({c.color})' if c.color else ''),
                  'qty': 1, 'amount': int(c.sold_price or 0)}]
        gifts = c.gift_label
        return self.render('admin/receipt.html', kind='Hoá đơn bán máy', ref=f'S{c.id:05d}',
                           customer=c.sold_to or '', phone=c.sold_phone or '',
                           items=items, deposit=int(c.deposit_amount or 0),
                           note=(('Tặng kèm: ' + gifts) if gifts else '') or c.sold_note or '',
                           **self._ctx())

    @expose('/accessory/<int:sid>')
    def accessory(self, sid):
        s = db.session.get(AccessorySale, sid)
        if not s:
            return 'Không tìm thấy giao dịch', 404
        items = [{'label': s.name or 'Phụ kiện', 'qty': s.quantity or 1, 'amount': s.revenue}]
        return self.render('admin/receipt.html', kind='Hoá đơn phụ kiện', ref=f'A{s.id:05d}',
                           customer=s.customer_name or '', phone=s.phone or '',
                           items=items, deposit=0, note=s.note or '', **self._ctx())


admin.add_view(BookingGridView(name='Lịch thuê (Sheet)', endpoint='bookinggrid'))
admin.add_view(CalendarView(name='Lịch (Calendar)',     endpoint='calendar'))
admin.add_view(CamerasView(name='Quản lý máy',          endpoint='cameras'))
admin.add_view(AccessoriesView(name='Phụ kiện',         endpoint='phukien'))
admin.add_view(CustomersView(name='Khách hàng',         endpoint='customers'))
admin.add_view(FinanceView(name='Tài chính',            endpoint='finance'))
admin.add_view(StoreCostView(name='Chi phí tiệm',       endpoint='chiphi'))
admin.add_view(LeadView(name='Khách quan tâm',          endpoint='leads'))
admin.add_view(EmployeesView(name='Nhân viên',          endpoint='employees'))
admin.add_view(ActivityView(name='Nhật ký',             endpoint='activity'))
admin.add_view(SheetView(name='Sổ tay',                 endpoint='sheet'))
admin.add_view(SettingsView(name='Cài đặt',             endpoint='settings'))
admin.add_view(MyView(name='Cá nhân',                   endpoint='my'))
admin.add_view(ReceiptView(name='Hoá đơn',              endpoint='receipt'))


# ── Schema migration (non-destructive) ─────────────────────────────────────────

def _auto_add_missing_columns():
    """Generic safety net: for every mapped table that already exists, add any column
    present in the model but missing in the DB (as a nullable column — safe on SQLite,
    never drops or rewrites data). Catches anything the explicit lists above miss."""
    insp = sa_inspect(db.engine)
    existing = set(insp.get_table_names())
    for table in db.metadata.sorted_tables:
        if table.name not in existing:
            continue                      # brand-new tables are created in full by create_all
        have = {c['name'] for c in insp.get_columns(table.name)}
        for col in table.columns:
            if col.name in have:
                continue
            try:
                coltype = col.type.compile(dialect=db.engine.dialect)
            except Exception:
                coltype = 'TEXT'
            default = ''
            d = getattr(col, 'default', None)
            if d is not None and getattr(d, 'is_scalar', False):
                v = d.arg
                if isinstance(v, bool):            default = f' DEFAULT {1 if v else 0}'
                elif isinstance(v, (int, float)):  default = f' DEFAULT {v}'
                elif isinstance(v, str):           default = " DEFAULT '{}'".format(v.replace("'", "''"))
            db.session.execute(text(f'ALTER TABLE {table.name} ADD COLUMN {col.name} {coltype}{default}'))
    db.session.commit()


def ensure_schema():
    """Add any new columns / tables without dropping existing data."""
    insp = sa_inspect(db.engine)
    if 'camera' in insp.get_table_names():
        cols = {c['name'] for c in insp.get_columns('camera')}
        new_cols = {
            'import_cost':   'INTEGER DEFAULT 0',
            'is_broken':     'BOOLEAN DEFAULT 0',
            'repair_cost':   'INTEGER DEFAULT 0',
            'is_sold':       'BOOLEAN DEFAULT 0',
            'sold_price':    'INTEGER DEFAULT 0',
            'deposit_amount':'INTEGER DEFAULT 0',
            'sold_date':     'DATETIME',
            'category':      "VARCHAR(40) DEFAULT ''",
            'reorder_point': 'INTEGER DEFAULT 0',
            'origin':        "VARCHAR(20) DEFAULT ''",
            'accessory':     "VARCHAR(120) DEFAULT ''",
            'color':         "VARCHAR(40) DEFAULT ''",
            'date_in':       'DATETIME',
            'sale_state':    "VARCHAR(20) DEFAULT 'stock'",
            'sold_to':       "VARCHAR(100) DEFAULT ''",
            'sold_phone':    "VARCHAR(20) DEFAULT ''",
            'sold_source':   "VARCHAR(30) DEFAULT ''",
            'gift_json':     "TEXT DEFAULT ''",
        }
        for name, ddl in new_cols.items():
            if name not in cols:
                db.session.execute(text(f'ALTER TABLE camera ADD COLUMN {name} {ddl}'))
        db.session.commit()

    db.create_all()               # creates brand-new tables (Appointment, Sheet…) in full
    _auto_add_missing_columns()   # backfill any column added to an already-existing table


# ── Auto-migrate on EVERY startup ──────────────────────────────────────────────
# Runs under Gunicorn too (the __main__ block below never executes there). This is
# non-destructive (only CREATE TABLE / ADD COLUMN) and never seeds, so the production
# database on the server keeps all its data and just gains the new schema.
with app.app_context():
    try:
        ensure_schema()
    except Exception as exc:      # never let a migration hiccup take the whole app down
        app.logger.warning(f'[schema] ensure_schema at startup failed: {exc}')


# ── Startup (local dev only: `python app.py`) ──────────────────────────────────
if __name__ == '__main__':
    with app.app_context():
        added  = seed_db(db, Camera)          # seeding is LOCAL ONLY — never runs on the server
        filled = backfill_costs(db, Camera)
        if added:
            print(f'[seed] Added {added} cameras to the database.')
        if filled:
            print(f'[seed] Backfilled import cost for {filled} sale cameras.')
    app.run(host='0.0.0.0', port=5000, debug=False)
